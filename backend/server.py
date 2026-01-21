from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Query
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timedelta
import pyxlsb
import tempfile
import re
import unicodedata

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Define Models
class LoginRequest(BaseModel):
    username: str
    password: str

class LoginResponse(BaseModel):
    success: bool
    message: str
    token: Optional[str] = None
    user: Optional[dict] = None

class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str

class ForgotPasswordRequest(BaseModel):
    username: str

class RutItem(BaseModel):
    rut_kod: str
    rut_aciklama: str
    ziyaret_sira: int
    musteri_kod: str
    musteri_unvan: str
    musteri_durum: str
    musteri_grup: str
    adres: Optional[str] = None

class RutTalepRequest(BaseModel):
    dst_name: str
    gun: str
    yeni_sira: List[dict]

class UserCreate(BaseModel):
    username: str
    password: str
    role: str  # admin, dsm, tte, dst
    name: str
    dst_name: Optional[str] = None  # For DST users

class BayiSummary(BaseModel):
    bayi_kodu: str
    bayi_unvani: str
    kapsam_durumu: Optional[str] = None
    tip: Optional[str] = None
    sinif: Optional[str] = None
    dst: Optional[str] = None
    tte: Optional[str] = None

class DashboardStats(BaseModel):
    aktif_bayi: int
    pasif_bayi: int

class DistributorTotals(BaseModel):
    # B22-AH22 verileri
    bayi_sayisi: Optional[float] = None
    aktif_bayi_sayisi: Optional[float] = None
    pasif_bayi_sayisi: Optional[float] = None
    aralik_hedef: Optional[float] = None
    aralik_satis: Optional[float] = None
    kalan_satis: Optional[float] = None
    hedef_basari_orani: Optional[float] = None
    tahsilat_hedef: Optional[float] = None
    tahsilat_tutari: Optional[float] = None
    ay_hedef_ziyaret: Optional[float] = None
    ziyaret_gerceklesen: Optional[float] = None
    drop_rate: Optional[float] = None
    basarili_satis: Optional[float] = None
    basarili_satis_yuzde: Optional[float] = None
    carili_bayi_sayisi: Optional[float] = None
    gun_0: Optional[float] = None
    gun_1: Optional[float] = None
    gun_2: Optional[float] = None
    gun_3: Optional[float] = None
    gun_4: Optional[float] = None
    gun_5: Optional[float] = None
    gun_6: Optional[float] = None
    gun_7: Optional[float] = None
    gun_8: Optional[float] = None
    gun_9: Optional[float] = None
    gun_10: Optional[float] = None
    gun_11: Optional[float] = None
    gun_12: Optional[float] = None
    gun_13: Optional[float] = None
    gun_14_uzeri: Optional[float] = None
    cari_toplam: Optional[float] = None
    loy_verilen_bayi_sayisi: Optional[float] = None
    loy_bayi_mahsuplasma_tutari: Optional[float] = None
    # Carili Kanal Bazlı Veriler (Konya Gün H2-H6, K2-K6)
    carili_piyasa: Optional[float] = None
    carili_yerel_zincir: Optional[float] = None
    carili_askeriye: Optional[float] = None
    carili_benzinlik: Optional[float] = None
    carili_geleneksel: Optional[float] = None
    # BV22-CZ22 verileri
    camel_hedef: Optional[float] = None
    winston_hedef: Optional[float] = None
    mcarlo_hedef: Optional[float] = None
    myo_camel_hedef: Optional[float] = None
    ld_hedef: Optional[float] = None
    toplam_hedef: Optional[float] = None
    kasa_hedef: Optional[float] = None
    hedef_das: Optional[float] = None
    camel_satis: Optional[float] = None
    winston_satis: Optional[float] = None
    mcarlo_satis: Optional[float] = None
    myo_camel_satis: Optional[float] = None
    ld_satis: Optional[float] = None
    toplam_satis: Optional[float] = None
    kasa_satis: Optional[float] = None
    gerc_das: Optional[float] = None
    bak_01: Optional[float] = None
    mar_02: Optional[float] = None
    bfe_03: Optional[float] = None
    kye_04: Optional[float] = None
    tek_05: Optional[float] = None
    ben_07: Optional[float] = None
    ask_08: Optional[float] = None
    czv_11: Optional[float] = None
    yznc_12: Optional[float] = None
    tut_14: Optional[float] = None
    tus_15: Optional[float] = None
    jti: Optional[float] = None
    pmi: Optional[float] = None
    bat: Optional[float] = None
    rut_say: Optional[float] = None
    # Q Line verileri
    qline_hedef: Optional[float] = None
    qline_2026_satis: Optional[float] = None
    qline_kalan: Optional[float] = None
    frekans_ort: Optional[float] = None
    # Son güncelleme zamanı
    son_guncelleme: Optional[str] = None

class BayiDetail(BaseModel):
    bayi_kodu: str
    bayi_unvani: str
    dst: Optional[str] = None
    tte: Optional[str] = None
    dsm: Optional[str] = None
    tip: Optional[str] = None
    panaroma_sinif: Optional[str] = None
    satisa_gore_sinif: Optional[str] = None
    kapsam_durumu: Optional[str] = None
    jti_stant: Optional[str] = None
    jti_stant_adet: Optional[float] = None
    camel_myo_stant: Optional[str] = None
    camel_myo_adet: Optional[float] = None
    pmi_stant: Optional[str] = None
    pmi_adet: Optional[float] = None
    bat_stant: Optional[str] = None
    bat_adet: Optional[float] = None
    loyalty_plan_2025: Optional[float] = None
    odenen_2025: Optional[float] = None
    # 2025 Monthly Sales
    ocak_2025: Optional[float] = None
    subat_2025: Optional[float] = None
    mart_2025: Optional[float] = None
    nisan_2025: Optional[float] = None
    mayis_2025: Optional[float] = None
    haziran_2025: Optional[float] = None
    temmuz_2025: Optional[float] = None
    agustos_2025: Optional[float] = None
    eylul_2025: Optional[float] = None
    ekim_2025: Optional[float] = None
    kasim_2025: Optional[float] = None
    aralik_2025: Optional[float] = None
    toplam_satis_2025: Optional[float] = None
    ortalama_2025: Optional[float] = None
    # 2024 Monthly Sales
    ocak_2024: Optional[float] = None
    subat_2024: Optional[float] = None
    mart_2024: Optional[float] = None
    nisan_2024: Optional[float] = None
    mayis_2024: Optional[float] = None
    haziran_2024: Optional[float] = None
    temmuz_2024: Optional[float] = None
    agustos_2024: Optional[float] = None
    eylul_2024: Optional[float] = None
    ekim_2024: Optional[float] = None
    kasim_2024: Optional[float] = None
    aralik_2024: Optional[float] = None
    toplam_satis_2024: Optional[float] = None
    ortalama_2024: Optional[float] = None
    # 2026 Monthly Sales
    ocak_2026: Optional[float] = None
    subat_2026: Optional[float] = None
    mart_2026: Optional[float] = None
    nisan_2026: Optional[float] = None
    mayis_2026: Optional[float] = None
    haziran_2026: Optional[float] = None
    temmuz_2026: Optional[float] = None
    agustos_2026: Optional[float] = None
    eylul_2026: Optional[float] = None
    ekim_2026: Optional[float] = None
    kasim_2026: Optional[float] = None
    aralik_2026: Optional[float] = None
    toplam_2026: Optional[float] = None
    ortalama_2026: Optional[float] = None
    # Development percentage
    gelisim_yuzdesi: Optional[float] = None
    # Debt
    borc_durumu: Optional[str] = None
    # Visit days
    ziyaret_gunleri: Optional[List[str]] = None

class Fatura(BaseModel):
    matbu_no: str
    tarih: str
    net_tutar: float
    bayi_kodu: str

class FaturaDetay(BaseModel):
    matbu_no: str
    urunler: List[Dict[str, Any]]
    toplam_miktar: float
    toplam_tutar: Optional[float] = 0

class Tahsilat(BaseModel):
    tahsilat_turu: str
    islem_tarihi: str
    tutar: float
    bayi_kodu: str

class PasifBayi(BaseModel):
    bayi_kodu: str
    bayi_unvani: str
    dst: Optional[str] = None
    tte: Optional[str] = None
    txtkapsam: Optional[str] = None

class DSTData(BaseModel):
    dst: str
    bayi_sayisi: Optional[float] = None
    aktif_bayi_sayisi: Optional[float] = None
    pasif_bayi_sayisi: Optional[float] = None
    aralik_hedef: Optional[float] = None
    aralik_satis: Optional[float] = None
    kalan_satis: Optional[float] = None
    hedef_basari_orani: Optional[float] = None
    tahsilat_hedef: Optional[float] = None
    tahsilat_tutari: Optional[float] = None
    ay_hedef_ziyaret: Optional[float] = None
    ziyaret_gerceklesen: Optional[float] = None
    drop_rate: Optional[float] = None
    basarili_satis: Optional[float] = None
    basarili_satis_yuzde: Optional[float] = None
    carili_bayi_sayisi: Optional[float] = None
    gun_0: Optional[float] = None
    gun_1: Optional[float] = None
    gun_2: Optional[float] = None
    gun_3: Optional[float] = None
    gun_4: Optional[float] = None
    gun_5: Optional[float] = None
    gun_6: Optional[float] = None
    gun_7: Optional[float] = None
    gun_8: Optional[float] = None
    gun_9: Optional[float] = None
    gun_10: Optional[float] = None
    gun_11: Optional[float] = None
    gun_12: Optional[float] = None
    gun_13: Optional[float] = None
    gun_14_uzeri: Optional[float] = None
    cari_toplam: Optional[float] = None
    loy_verilen_bayi_sayisi: Optional[float] = None
    loy_bayi_mahsuplasma_tutari: Optional[float] = None
    # SKU Satışları
    skt_camel_yellow_100: Optional[float] = None
    camel_brown: Optional[float] = None
    camel_black: Optional[float] = None
    camel_white: Optional[float] = None
    camel_yellow_sp: Optional[float] = None
    camel_yellow: Optional[float] = None
    camel_deep_blue_long: Optional[float] = None
    camel_deep_blue: Optional[float] = None
    camel_yellow_long: Optional[float] = None
    camel_slender_blue: Optional[float] = None
    dp_camel_slender_blueline: Optional[float] = None
    camel_slender_gray: Optional[float] = None
    dp_camel_slender_grayline: Optional[float] = None
    winston_red_long: Optional[float] = None
    winston_red: Optional[float] = None
    winston_blue_long: Optional[float] = None
    winston_blue: Optional[float] = None
    winston_gray: Optional[float] = None
    winston_slims_blue: Optional[float] = None
    winston_slims_gray: Optional[float] = None
    winston_slims_q_line: Optional[float] = None
    winston_xsence_black: Optional[float] = None
    winston_xsence_gray: Optional[float] = None
    winston_dark_blue_long: Optional[float] = None
    winston_dark_blue: Optional[float] = None
    winston_deep_blue: Optional[float] = None
    winston_slender_blue_long: Optional[float] = None
    winston_slender_blue: Optional[float] = None
    winston_slender_gray: Optional[float] = None
    winston_slender_dark_blue: Optional[float] = None
    winston_slender_q_line: Optional[float] = None
    monte_carlo_red: Optional[float] = None
    monte_carlo_dark_blue_long: Optional[float] = None
    monte_carlo_dark_blue: Optional[float] = None
    monte_carlo_slender_dark_blue: Optional[float] = None
    ld_slims: Optional[float] = None
    ld_blue_long: Optional[float] = None
    ld_blue: Optional[float] = None
    toplam_gun_sku: Optional[float] = None
    # Marka Toplamları
    camel_toplam: Optional[float] = None
    winston_toplam: Optional[float] = None
    mcarlo_toplam: Optional[float] = None
    myo_camel: Optional[float] = None
    ld_toplam: Optional[float] = None
    toplam: Optional[float] = None
    kasa: Optional[float] = None
    hedef_das: Optional[float] = None
    # Gerçekleşen
    camel_gerc: Optional[float] = None
    winston_gerc: Optional[float] = None
    mcarlo_gerc: Optional[float] = None
    myo_camel_gerc: Optional[float] = None
    ld_gerc: Optional[float] = None
    toplam_gerc: Optional[float] = None
    kasa_gerc: Optional[float] = None
    gerc_das: Optional[float] = None
    # Kanal Bazlı
    bak_01: Optional[float] = None
    mar_02: Optional[float] = None
    bfe_03: Optional[float] = None
    kye_04: Optional[float] = None
    tek_05: Optional[float] = None
    ben_07: Optional[float] = None
    ask_08: Optional[float] = None
    czv_11: Optional[float] = None
    yznc_12: Optional[float] = None
    tut_14: Optional[float] = None
    tus_15: Optional[float] = None
    # Stand Sayıları
    jti: Optional[float] = None
    pmi: Optional[float] = None
    bat: Optional[float] = None
    rut_say: Optional[float] = None
    # Yıllık SKU
    w_dark_blue_ks: Optional[float] = None
    w_slender_blue_ks: Optional[float] = None
    w_dark_blue_long: Optional[float] = None
    mcarlo_slender_dark_blue_yil: Optional[float] = None
    w_slim_blue: Optional[float] = None
    w_blue_ks: Optional[float] = None
    w_slender_blue_long: Optional[float] = None
    camel_slender_blue_yil: Optional[float] = None
    mcarlo_dark_blue_ks: Optional[float] = None
    mcarlo_dark_blue_long_yil: Optional[float] = None
    w_slender_q_line_2025: Optional[float] = None
    w_slender_q_line_2026: Optional[float] = None
    frekans_ort: Optional[float] = None

# DSM Team Data Model
class DSMTeamData(BaseModel):
    team_name: str
    dsm_name: str
    dst_list: List[str]
    bayi_sayisi: Optional[float] = None
    aktif_bayi_sayisi: Optional[float] = None
    pasif_bayi_sayisi: Optional[float] = None
    aralik_hedef: Optional[float] = None
    aralik_satis: Optional[float] = None
    kalan_satis: Optional[float] = None
    hedef_basari_orani: Optional[float] = None
    tahsilat_hedef: Optional[float] = None
    tahsilat_tutari: Optional[float] = None
    ay_hedef_ziyaret: Optional[float] = None
    ziyaret_gerceklesen: Optional[float] = None
    drop_rate: Optional[float] = None
    basarili_satis: Optional[float] = None
    basarili_satis_yuzde: Optional[float] = None
    carili_bayi_sayisi: Optional[float] = None
    gun_0: Optional[float] = None
    gun_1: Optional[float] = None
    gun_2: Optional[float] = None
    gun_3: Optional[float] = None
    gun_4: Optional[float] = None
    gun_5: Optional[float] = None
    gun_6: Optional[float] = None
    gun_7: Optional[float] = None
    gun_8: Optional[float] = None
    gun_9: Optional[float] = None
    gun_10: Optional[float] = None
    gun_11: Optional[float] = None
    gun_12: Optional[float] = None
    gun_13: Optional[float] = None
    gun_14_uzeri: Optional[float] = None
    cari_toplam: Optional[float] = None
    # Son Gün SKU
    toplam_gun_sku: Optional[float] = None
    # Hedefler
    camel_toplam: Optional[float] = None
    winston_toplam: Optional[float] = None
    mcarlo_toplam: Optional[float] = None
    myo_camel: Optional[float] = None
    ld_toplam: Optional[float] = None
    toplam: Optional[float] = None
    kasa: Optional[float] = None
    hedef_das: Optional[float] = None
    # Satışlar
    camel_gerc: Optional[float] = None
    winston_gerc: Optional[float] = None
    mcarlo_gerc: Optional[float] = None
    myo_camel_gerc: Optional[float] = None
    ld_gerc: Optional[float] = None
    toplam_gerc: Optional[float] = None
    kasa_gerc: Optional[float] = None
    gerc_das: Optional[float] = None
    # Bayi Tipleri
    bak_01: Optional[float] = None
    mar_02: Optional[float] = None
    bfe_03: Optional[float] = None
    kye_04: Optional[float] = None
    tek_05: Optional[float] = None
    ben_07: Optional[float] = None
    ask_08: Optional[float] = None
    czv_11: Optional[float] = None
    yznc_12: Optional[float] = None
    tut_14: Optional[float] = None
    tus_15: Optional[float] = None
    jti: Optional[float] = None
    pmi: Optional[float] = None
    bat: Optional[float] = None
    rut_say: Optional[float] = None
    # İlk 10 SKU
    w_dark_blue_ks: Optional[float] = None
    w_slender_blue_ks: Optional[float] = None
    w_dark_blue_long: Optional[float] = None
    mcarlo_slender_dark_blue_yil: Optional[float] = None
    w_slim_blue: Optional[float] = None
    w_blue_ks: Optional[float] = None
    w_slender_blue_long: Optional[float] = None
    camel_slender_blue_yil: Optional[float] = None
    mcarlo_dark_blue_ks: Optional[float] = None
    mcarlo_dark_blue_long_yil: Optional[float] = None
    w_slender_q_line_2025: Optional[float] = None
    w_slender_q_line_2026: Optional[float] = None
    frekans_ort: Optional[float] = None
    # Aylık Satışlar (from AÜ BAYİ LİST)
    ocak_2025: Optional[float] = None
    subat_2025: Optional[float] = None
    mart_2025: Optional[float] = None
    nisan_2025: Optional[float] = None
    mayis_2025: Optional[float] = None
    haziran_2025: Optional[float] = None
    temmuz_2025: Optional[float] = None
    agustos_2025: Optional[float] = None
    eylul_2025: Optional[float] = None
    ekim_2025: Optional[float] = None
    kasim_2025: Optional[float] = None
    aralik_2025: Optional[float] = None
    toplam_satis_2025: Optional[float] = None

# TTE Data Model
class TTEData(BaseModel):
    tte_name: str
    bayi_sayisi: Optional[float] = None
    aktif_bayi_sayisi: Optional[float] = None
    pasif_bayi_sayisi: Optional[float] = None
    jti: Optional[float] = None
    jti_stand: Optional[float] = None
    pmi: Optional[float] = None
    pmi_stand: Optional[float] = None
    bat: Optional[float] = None
    bat_stand: Optional[float] = None
    sinif_a: Optional[float] = None
    sinif_a_plus: Optional[float] = None
    sinif_b: Optional[float] = None
    sinif_c: Optional[float] = None
    sinif_d: Optional[float] = None
    sinif_e: Optional[float] = None
    sinif_e_minus: Optional[float] = None

class CariBayi(BaseModel):
    bayi_kodu: str
    unvan: str
    dst: str
    dsm: str
    tip: Optional[str] = None
    sinif: Optional[str] = None
    musteri_bakiyesi: Optional[float] = None
    gun_0: Optional[float] = None
    gun_1: Optional[float] = None
    gun_2: Optional[float] = None
    gun_3: Optional[float] = None
    gun_4: Optional[float] = None
    gun_5: Optional[float] = None
    gun_6: Optional[float] = None
    gun_7: Optional[float] = None
    gun_8: Optional[float] = None
    gun_9: Optional[float] = None
    gun_10: Optional[float] = None
    gun_11: Optional[float] = None
    gun_12: Optional[float] = None
    gun_13: Optional[float] = None
    gun_14_uzeri: Optional[float] = None

# Helper function to convert Turkish characters for case-insensitive search
def turkish_to_ascii(text: str) -> str:
    """Convert Turkish special characters to ASCII equivalents"""
    if not text:
        return ""
    # Extended map including all Turkish characters
    tr_map = {
        'ç': 'c', 'Ç': 'c',
        'ğ': 'g', 'Ğ': 'g',
        'ı': 'i', 'I': 'i', 'İ': 'i', 'i': 'i',
        'ö': 'o', 'Ö': 'o',
        'ş': 's', 'Ş': 's',
        'ü': 'u', 'Ü': 'u',
        # Additional Turkish chars that might appear
        '\u0130': 'i',  # İ (Turkish capital I with dot)
        '\u0131': 'i',  # ı (Turkish lowercase dotless i)
    }
    result = text.lower()
    for tr_char, ascii_char in tr_map.items():
        result = result.replace(tr_char, ascii_char)
    # Also normalize unicode
    import unicodedata
    result = unicodedata.normalize('NFKD', result)
    result = ''.join(c for c in result if not unicodedata.combining(c))
    return result

# Helper to parse Excel date serial
def excel_date_to_str(serial):
    if serial is None:
        return None
    try:
        if isinstance(serial, str):
            return serial
        base_date = datetime(1899, 12, 30)
        date = base_date + timedelta(days=int(serial))
        return date.strftime('%d/%m/%Y')
    except:
        return str(serial) if serial else None

def safe_float(value):
    if value is None:
        return 0.0
    try:
        return float(value)
    except:
        return 0.0

def safe_str(value):
    if value is None:
        return None
    return str(value).strip() if value else None

# DST listesi ve şifreleri
DST_USERS = {
    "dst1": {"name": "KEMAL BANİ", "password": "dst1konya"},
    "dst2": {"name": "COŞKUN ÇİMEN", "password": "dst2konya"},
    "dst3": {"name": "MUSTAFA KAĞAN KAYA", "password": "dst3konya"},
    "dst4": {"name": "MUSTAFA HARMANCI", "password": "dst4konya"},
    "dst5": {"name": "KAZIM KARABEKİR ÖRAN", "password": "dst5konya"},
    "dst6": {"name": "TUNAHAN IŞILAK", "password": "dst6konya"},
    "dst7": {"name": "MEVLÜT ŞEKER", "password": "dst7konya"},
    "dst8": {"name": "TAHİR UÇAR", "password": "dst8konya"},
    "dst9": {"name": "YASİN TUĞRA DAĞLI", "password": "dst9konya"},
    "dst10": {"name": "HÜSEYİN AYHAN AKMAN", "password": "dst10konya"},
    "dst11": {"name": "MUSTAFA USLU", "password": "dst11konya"},
    "dst12": {"name": "HASAN ALİ AKDAĞ", "password": "dst12konya"},
    "dst13": {"name": "AHMET GÖKMEN", "password": "dst13konya"},
    "dst14": {"name": "LÜTFİ UYSAL", "password": "dst14konya"},
    "dst15": {"name": "ŞERAFETTİN BÜYÜKTAŞDELEN", "password": "dst15konya"},
    "dst16": {"name": "BURAK KÜÇÜKŞANTÜRK", "password": "dst16konya"},
    "dst17": {"name": "YASİN AVCI", "password": "dst17konya"},
    "dst18": {"name": "MUSTAFA İBİŞ", "password": "dst18konya"},
}

# Login endpoint
@api_router.post("/login", response_model=LoginResponse)
async def login(request: LoginRequest):
    username = request.username.lower().strip()
    password = request.password.strip()
    
    # Check database first for user with changed password
    db_user = await db.users.find_one({"username": username})
    if db_user:
        if db_user.get("password") == password:
            return LoginResponse(
                success=True, 
                message="Giriş başarılı", 
                token=f"{username}-token",
                user={
                    "username": username,
                    "name": db_user.get("name"),
                    "role": db_user.get("role"),
                    "dst_name": db_user.get("dst_name")
                }
            )
        return LoginResponse(success=False, message="Şifre hatalı")
    
    # Admin authentication
    if username == "admin" and password == "admin123":
        return LoginResponse(
            success=True, 
            message="Giriş başarılı", 
            token="admin-token",
            user={"username": "admin", "name": "Administrator", "role": "admin"}
        )
    
    # DST authentication
    if username in DST_USERS:
        dst_info = DST_USERS[username]
        if password == dst_info["password"]:
            return LoginResponse(
                success=True, 
                message="Giriş başarılı", 
                token=f"{username}-token",
                user={
                    "username": username,
                    "name": dst_info["name"],
                    "role": "dst",
                    "dst_name": dst_info["name"]
                }
            )
        return LoginResponse(success=False, message="Şifre hatalı")
    
    return LoginResponse(success=False, message="Kullanıcı adı veya şifre hatalı")

# Change password endpoint
@api_router.post("/change-password")
async def change_password(request: ChangePasswordRequest, username: str = Query(...)):
    username = username.lower().strip()
    
    # Verify old password first
    db_user = await db.users.find_one({"username": username})
    
    if db_user:
        if db_user.get("password") != request.old_password:
            return {"success": False, "message": "Mevcut şifre hatalı"}
    else:
        # Check default passwords
        if username == "admin":
            if request.old_password != "admin123":
                return {"success": False, "message": "Mevcut şifre hatalı"}
            role = "admin"
            name = "Administrator"
            dst_name = None
        elif username in DST_USERS:
            if request.old_password != DST_USERS[username]["password"]:
                return {"success": False, "message": "Mevcut şifre hatalı"}
            role = "dst"
            name = DST_USERS[username]["name"]
            dst_name = DST_USERS[username]["name"]
        else:
            return {"success": False, "message": "Kullanıcı bulunamadı"}
    
    # Update or create user with new password
    await db.users.update_one(
        {"username": username},
        {"$set": {
            "username": username,
            "password": request.new_password,
            "role": db_user.get("role") if db_user else role,
            "name": db_user.get("name") if db_user else name,
            "dst_name": db_user.get("dst_name") if db_user else dst_name
        }},
        upsert=True
    )
    
    return {"success": True, "message": "Şifre başarıyla değiştirildi"}

# Forgot password endpoint
@api_router.post("/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    username = request.username.lower().strip()
    
    # Check if user exists
    user_exists = False
    user_name = None
    
    if username == "admin":
        user_exists = True
        user_name = "Administrator"
    elif username in DST_USERS:
        user_exists = True
        user_name = DST_USERS[username]["name"]
    else:
        db_user = await db.users.find_one({"username": username})
        if db_user:
            user_exists = True
            user_name = db_user.get("name")
    
    if not user_exists:
        return {"success": False, "message": "Kullanıcı bulunamadı"}
    
    # Generate new password
    import random
    import string
    new_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    
    # Update password in database
    await db.users.update_one(
        {"username": username},
        {"$set": {"password": new_password, "name": user_name}},
        upsert=True
    )
    
    # Log the password reset request (in real app, send email)
    logger.info(f"Password reset for {username} ({user_name}). New password: {new_password}")
    logger.info(f"Email should be sent to: operator@aydinunluer.com.tr")
    
    # In production, send email here
    # For now, just return success
    return {
        "success": True, 
        "message": "Yeni şifreniz operator@aydinunluer.com.tr adresine gönderildi"
    }

# Get all users (admin only)
@api_router.get("/users")
async def get_users():
    users = []
    # Add admin
    users.append({"username": "admin", "name": "Administrator", "role": "admin"})
    # Add DST users
    for username, info in DST_USERS.items():
        users.append({
            "username": username,
            "name": info["name"],
            "role": "dst",
            "dst_name": info["name"]
        })
    return users

# Dashboard stats
@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats():
    try:
        aktif = await db.stand_raporu.count_documents({"bayi_durumu": "Aktif"})
        pasif = await db.stand_raporu.count_documents({"bayi_durumu": "Pasif"})
        return DashboardStats(aktif_bayi=aktif, pasif_bayi=pasif)
    except Exception as e:
        logger.error(f"Error getting dashboard stats: {e}")
        return DashboardStats(aktif_bayi=0, pasif_bayi=0)

# Distributor totals endpoint
@api_router.get("/distributor-totals", response_model=DistributorTotals)
async def get_distributor_totals():
    try:
        totals = await db.distributor_totals.find_one({"type": "totals"})
        if totals:
            return DistributorTotals(**{k: v for k, v in totals.items() if k != "_id" and k != "type"})
        return DistributorTotals()
    except Exception as e:
        logger.error(f"Error getting distributor totals: {e}")
        return DistributorTotals()

# Pasif Bayiler listesi
@api_router.get("/pasif-bayiler", response_model=List[PasifBayi])
async def get_pasif_bayiler():
    try:
        # Get all passive dealers from stand_raporu
        pasif_list = await db.stand_raporu.find({"bayi_durumu": "Pasif"}).to_list(1000)
        
        result = []
        for p in pasif_list:
            bayi_kodu = p.get("bayi_kodu", "")
            
            # Get additional info from bayiler collection
            bayi = await db.bayiler.find_one({"bayi_kodu": bayi_kodu})
            
            result.append(PasifBayi(
                bayi_kodu=bayi_kodu,
                bayi_unvani=bayi.get("bayi_unvani", "") if bayi else "",
                dst=bayi.get("dst") if bayi else None,
                tte=bayi.get("tte") if bayi else None,
                txtkapsam=p.get("txtkapsam")
            ))
        
        # Sort by bayi_unvani
        result.sort(key=lambda x: x.bayi_unvani or "")
        return result
    except Exception as e:
        logger.error(f"Error getting pasif bayiler: {e}")
        return []

# DST Data listesi
@api_router.get("/dst-data", response_model=List[DSTData])
async def get_dst_data():
    try:
        dst_list = await db.dst_data.find({}).to_list(100)
        result = []
        for d in dst_list:
            result.append(DSTData(
                dst=d.get("dst", ""),
                bayi_sayisi=d.get("bayi_sayisi"),
                aktif_bayi_sayisi=d.get("aktif_bayi_sayisi"),
                pasif_bayi_sayisi=d.get("pasif_bayi_sayisi"),
                aralik_hedef=d.get("aralik_hedef"),
                aralik_satis=d.get("aralik_satis"),
                kalan_satis=d.get("kalan_satis"),
                hedef_basari_orani=d.get("hedef_basari_orani"),
                tahsilat_hedef=d.get("tahsilat_hedef"),
                tahsilat_tutari=d.get("tahsilat_tutari"),
                ay_hedef_ziyaret=d.get("ay_hedef_ziyaret"),
                ziyaret_gerceklesen=d.get("ziyaret_gerceklesen"),
                drop_rate=d.get("drop_rate"),
                basarili_satis=d.get("basarili_satis"),
                basarili_satis_yuzde=d.get("basarili_satis_yuzde"),
                carili_bayi_sayisi=d.get("carili_bayi_sayisi"),
                gun_0=d.get("gun_0"),
                gun_1=d.get("gun_1"),
                gun_2=d.get("gun_2"),
                gun_3=d.get("gun_3"),
                gun_4=d.get("gun_4"),
                gun_5=d.get("gun_5"),
                gun_6=d.get("gun_6"),
                gun_7=d.get("gun_7"),
                gun_8=d.get("gun_8"),
                gun_9=d.get("gun_9"),
                gun_10=d.get("gun_10"),
                gun_11=d.get("gun_11"),
                gun_12=d.get("gun_12"),
                gun_13=d.get("gun_13"),
                gun_14_uzeri=d.get("gun_14_uzeri"),
                cari_toplam=d.get("cari_toplam"),
                loy_verilen_bayi_sayisi=d.get("loy_verilen_bayi_sayisi"),
                loy_bayi_mahsuplasma_tutari=d.get("loy_bayi_mahsuplasma_tutari"),
                skt_camel_yellow_100=d.get("skt_camel_yellow_100"),
                camel_brown=d.get("camel_brown"),
                camel_black=d.get("camel_black"),
                camel_white=d.get("camel_white"),
                camel_yellow_sp=d.get("camel_yellow_sp"),
                camel_yellow=d.get("camel_yellow"),
                camel_deep_blue_long=d.get("camel_deep_blue_long"),
                camel_deep_blue=d.get("camel_deep_blue"),
                camel_yellow_long=d.get("camel_yellow_long"),
                camel_slender_blue=d.get("camel_slender_blue"),
                dp_camel_slender_blueline=d.get("dp_camel_slender_blueline"),
                camel_slender_gray=d.get("camel_slender_gray"),
                dp_camel_slender_grayline=d.get("dp_camel_slender_grayline"),
                winston_red_long=d.get("winston_red_long"),
                winston_red=d.get("winston_red"),
                winston_blue_long=d.get("winston_blue_long"),
                winston_blue=d.get("winston_blue"),
                winston_gray=d.get("winston_gray"),
                winston_slims_blue=d.get("winston_slims_blue"),
                winston_slims_gray=d.get("winston_slims_gray"),
                winston_slims_q_line=d.get("winston_slims_q_line"),
                winston_xsence_black=d.get("winston_xsence_black"),
                winston_xsence_gray=d.get("winston_xsence_gray"),
                winston_dark_blue_long=d.get("winston_dark_blue_long"),
                winston_dark_blue=d.get("winston_dark_blue"),
                winston_deep_blue=d.get("winston_deep_blue"),
                winston_slender_blue_long=d.get("winston_slender_blue_long"),
                winston_slender_blue=d.get("winston_slender_blue"),
                winston_slender_gray=d.get("winston_slender_gray"),
                winston_slender_dark_blue=d.get("winston_slender_dark_blue"),
                winston_slender_q_line=d.get("winston_slender_q_line"),
                monte_carlo_red=d.get("monte_carlo_red"),
                monte_carlo_dark_blue_long=d.get("monte_carlo_dark_blue_long"),
                monte_carlo_dark_blue=d.get("monte_carlo_dark_blue"),
                monte_carlo_slender_dark_blue=d.get("monte_carlo_slender_dark_blue"),
                ld_slims=d.get("ld_slims"),
                ld_blue_long=d.get("ld_blue_long"),
                ld_blue=d.get("ld_blue"),
                toplam_gun_sku=d.get("toplam_gun_sku"),
                camel_toplam=d.get("camel_toplam"),
                winston_toplam=d.get("winston_toplam"),
                mcarlo_toplam=d.get("mcarlo_toplam"),
                myo_camel=d.get("myo_camel"),
                ld_toplam=d.get("ld_toplam"),
                toplam=d.get("toplam"),
                kasa=d.get("kasa"),
                hedef_das=d.get("hedef_das"),
                camel_gerc=d.get("camel_gerc"),
                winston_gerc=d.get("winston_gerc"),
                mcarlo_gerc=d.get("mcarlo_gerc"),
                myo_camel_gerc=d.get("myo_camel_gerc"),
                ld_gerc=d.get("ld_gerc"),
                toplam_gerc=d.get("toplam_gerc"),
                kasa_gerc=d.get("kasa_gerc"),
                gerc_das=d.get("gerc_das"),
                bak_01=d.get("bak_01"),
                mar_02=d.get("mar_02"),
                bfe_03=d.get("bfe_03"),
                kye_04=d.get("kye_04"),
                tek_05=d.get("tek_05"),
                ben_07=d.get("ben_07"),
                ask_08=d.get("ask_08"),
                czv_11=d.get("czv_11"),
                yznc_12=d.get("yznc_12"),
                tut_14=d.get("tut_14"),
                tus_15=d.get("tus_15"),
                jti=d.get("jti"),
                pmi=d.get("pmi"),
                bat=d.get("bat"),
                rut_say=d.get("rut_say"),
                w_dark_blue_ks=d.get("w_dark_blue_ks"),
                w_slender_blue_ks=d.get("w_slender_blue_ks"),
                w_dark_blue_long=d.get("w_dark_blue_long"),
                mcarlo_slender_dark_blue_yil=d.get("mcarlo_slender_dark_blue_yil"),
                w_slim_blue=d.get("w_slim_blue"),
                w_blue_ks=d.get("w_blue_ks"),
                w_slender_blue_long=d.get("w_slender_blue_long"),
                camel_slender_blue_yil=d.get("camel_slender_blue_yil"),
                mcarlo_dark_blue_ks=d.get("mcarlo_dark_blue_ks"),
                mcarlo_dark_blue_long_yil=d.get("mcarlo_dark_blue_long_yil"),
                w_slender_q_line_2025=d.get("w_slender_q_line_2025"),
                w_slender_q_line_2026=d.get("w_slender_q_line_2026"),
                frekans_ort=d.get("frekans_ort"),
            ))
        return result
    except Exception as e:
        logger.error(f"Error getting DST data: {e}")
        return []

# DSM Team Data
@api_router.get("/dsm-teams")
async def get_dsm_teams():
    try:
        dsm_teams = await db.dsm_teams.find({}).to_list(10)
        # Remove MongoDB _id for JSON serialization
        for team in dsm_teams:
            team.pop('_id', None)
        return dsm_teams
    except Exception as e:
        logger.error(f"Error getting DSM teams: {e}")
        return []

# TTE Data
@api_router.get("/tte-data")
async def get_tte_data():
    try:
        tte_list = await db.tte_data.find({}).to_list(10)
        # Remove MongoDB _id for JSON serialization
        for tte in tte_list:
            tte.pop('_id', None)
        return tte_list
    except Exception as e:
        logger.error(f"Error getting TTE data: {e}")
        return []

# Cari Bayiler by DST and day
@api_router.get("/cari-bayiler/{dst}")
async def get_cari_bayiler(dst: str, gun: str = Query(default="toplam", description="Gun filtresi: 0-14, 14_uzeri, toplam")):
    try:
        dst_decoded = dst
        
        # Find all dealers for this DST with debt
        query = {"dst": dst_decoded}
        
        bayiler = await db.konya_gun.find(query).to_list(500)
        result = []
        
        for b in bayiler:
            # Filter based on gun parameter
            if gun == "toplam":
                # Show all dealers with any debt (musteri_bakiyesi > 0)
                if b.get("musteri_bakiyesi", 0) > 0:
                    result.append({
                        "bayi_kodu": str(b.get("bayi_kodu", "")),
                        "unvan": b.get("unvan", ""),
                        "dst": b.get("dst", ""),
                        "dsm": b.get("dsm", ""),
                        "tip": b.get("tip"),
                        "sinif": b.get("sinif"),
                        "musteri_bakiyesi": b.get("musteri_bakiyesi", 0),
                        "gun_deger": b.get("musteri_bakiyesi", 0)
                    })
            else:
                # Map gun parameter to field name
                gun_field_map = {
                    "0": "gun_0", "1": "gun_1", "2": "gun_2", "3": "gun_3",
                    "4": "gun_4", "5": "gun_5", "6": "gun_6", "7": "gun_7",
                    "8": "gun_8", "9": "gun_9", "10": "gun_10", "11": "gun_11",
                    "12": "gun_12", "13": "gun_13", "14_uzeri": "gun_14_uzeri"
                }
                field = gun_field_map.get(gun, "musteri_bakiyesi")
                value = b.get(field, 0)
                
                if value and value > 0:
                    result.append({
                        "bayi_kodu": str(b.get("bayi_kodu", "")),
                        "unvan": b.get("unvan", ""),
                        "dst": b.get("dst", ""),
                        "dsm": b.get("dsm", ""),
                        "tip": b.get("tip"),
                        "sinif": b.get("sinif"),
                        "musteri_bakiyesi": b.get("musteri_bakiyesi", 0),
                        "gun_deger": value
                    })
        
        # Sort by gun_deger descending
        result.sort(key=lambda x: x.get("gun_deger", 0), reverse=True)
        return result
    except Exception as e:
        logger.error(f"Error getting cari bayiler: {e}")
        return []

# Pasif Bayiler by DST
@api_router.get("/pasif-bayiler-dst/{dst}")
async def get_pasif_bayiler_dst(dst: str):
    try:
        # Get passive dealers from stand_raporu filtered by DST
        pipeline = [
            {"$match": {"bayi_durumu": "Pasif"}},
            {"$lookup": {
                "from": "bayiler",
                "localField": "bayi_kodu",
                "foreignField": "bayi_kodu",
                "as": "bayi_info"
            }},
            {"$unwind": {"path": "$bayi_info", "preserveNullAndEmptyArrays": True}},
            {"$match": {"bayi_info.dst": dst}}
        ]
        
        pasif_list = await db.stand_raporu.aggregate(pipeline).to_list(500)
        
        result = []
        for p in pasif_list:
            bayi_info = p.get("bayi_info", {})
            result.append({
                "bayi_kodu": p.get("bayi_kodu", ""),
                "bayi_unvani": bayi_info.get("bayi_unvani", ""),
                "dst": bayi_info.get("dst"),
                "tte": bayi_info.get("tte"),
                "txtkapsam": p.get("txtkapsam")
            })
        
        result.sort(key=lambda x: x.get("bayi_unvani", "") or "")
        return result
    except Exception as e:
        logger.error(f"Error getting pasif bayiler by DST: {e}")
        return []

# Pasif Bayiler by DSM
@api_router.get("/pasif-bayiler-dsm/{dsm}")
async def get_pasif_bayiler_dsm(dsm: str):
    try:
        # Get passive dealers from stand_raporu filtered by DSM
        pipeline = [
            {"$match": {"bayi_durumu": "Pasif"}},
            {"$lookup": {
                "from": "bayiler",
                "localField": "bayi_kodu",
                "foreignField": "bayi_kodu",
                "as": "bayi_info"
            }},
            {"$unwind": {"path": "$bayi_info", "preserveNullAndEmptyArrays": True}},
            {"$match": {"bayi_info.dsm": dsm}}
        ]
        
        pasif_list = await db.stand_raporu.aggregate(pipeline).to_list(500)
        
        result = []
        for p in pasif_list:
            bayi_info = p.get("bayi_info", {})
            result.append({
                "bayi_kodu": p.get("bayi_kodu", ""),
                "bayi_unvani": bayi_info.get("bayi_unvani", ""),
                "dst": bayi_info.get("dst"),
                "tte": bayi_info.get("tte"),
                "txtkapsam": p.get("txtkapsam")
            })
        
        result.sort(key=lambda x: x.get("bayi_unvani", "") or "")
        return result
    except Exception as e:
        logger.error(f"Error getting pasif bayiler by DSM: {e}")
        return []

# Pasif Bayiler by TTE
@api_router.get("/pasif-bayiler-tte/{tte}")
async def get_pasif_bayiler_tte(tte: str):
    try:
        # Get passive dealers from stand_raporu filtered by TTE
        # Use turkish_to_ascii for comparison to handle Turkish character differences
        tte_ascii = turkish_to_ascii(tte)
        
        pasif_list = await db.stand_raporu.find({
            "bayi_durumu": "Pasif"
        }).to_list(1000)
        
        result = []
        for p in pasif_list:
            db_tte = p.get("tte") or ""
            db_tte_ascii = turkish_to_ascii(db_tte)
            if db_tte_ascii == tte_ascii:
                result.append({
                    "bayi_kodu": p.get("bayi_kodu", ""),
                    "bayi_unvani": p.get("bayi_unvani", ""),
                    "dst": p.get("dst"),
                    "tte": p.get("tte"),
                    "txtkapsam": p.get("txtkapsam")
                })
        
        result.sort(key=lambda x: x.get("bayi_unvani", "") or "")
        return result
    except Exception as e:
        logger.error(f"Error getting pasif bayiler by TTE: {e}")
        return []

# Tüm distribütör için carili bayiler (gün bazlı)
@api_router.get("/cari-bayiler-tumu")
async def get_cari_bayiler_tumu(gun: str = Query(..., description="Gün değeri: 0, 1, 2, ... 14_uzeri, toplam")):
    try:
        # Map gun parameter to field name
        gun_mapping = {
            "0": "gun_0", "1": "gun_1", "2": "gun_2", "3": "gun_3",
            "4": "gun_4", "5": "gun_5", "6": "gun_6", "7": "gun_7",
            "8": "gun_8", "9": "gun_9", "10": "gun_10", "11": "gun_11",
            "12": "gun_12", "13": "gun_13", "14_uzeri": "gun_14_uzeri",
            "toplam": "musteri_bakiyesi"
        }
        
        field_name = gun_mapping.get(gun, "musteri_bakiyesi")
        
        # Get all records from konya_gun where the specified day field > 0
        query = {field_name: {"$gt": 0}}
        
        records = await db.konya_gun.find(query).to_list(1000)
        
        result = []
        for r in records:
            gun_value = r.get(field_name, 0) or 0
            if gun_value > 0:
                result.append({
                    "bayi_kodu": r.get("bayi_kodu", ""),
                    "unvan": r.get("unvan", ""),
                    "dst": r.get("dst", ""),
                    "dsm": r.get("dsm", ""),
                    "tip": r.get("tip", ""),
                    "sinif": r.get("sinif", ""),
                    "musteri_bakiyesi": r.get("musteri_bakiyesi", 0),
                    "gun_deger": gun_value
                })
        
        # Sort by gun_deger descending
        result.sort(key=lambda x: x.get("gun_deger", 0), reverse=True)
        return result
    except Exception as e:
        logger.error(f"Error getting all cari bayiler: {e}")
        return []

# DSM için carili bayiler (gün bazlı)
@api_router.get("/cari-bayiler-dsm/{dsm}")
async def get_cari_bayiler_dsm(dsm: str, gun: str = Query(..., description="Gün değeri")):
    try:
        gun_mapping = {
            "0": "gun_0", "1": "gun_1", "2": "gun_2", "3": "gun_3",
            "4": "gun_4", "5": "gun_5", "6": "gun_6", "7": "gun_7",
            "8": "gun_8", "9": "gun_9", "10": "gun_10", "11": "gun_11",
            "12": "gun_12", "13": "gun_13", "14_uzeri": "gun_14_uzeri",
            "toplam": "musteri_bakiyesi"
        }
        
        field_name = gun_mapping.get(gun, "musteri_bakiyesi")
        dsm_ascii = turkish_to_ascii(dsm)
        
        # Get all records and filter by DSM
        all_records = await db.konya_gun.find().to_list(1000)
        
        result = []
        for r in all_records:
            db_dsm = r.get("dsm", "") or ""
            db_dsm_ascii = turkish_to_ascii(db_dsm)
            
            if db_dsm_ascii == dsm_ascii:
                gun_value = r.get(field_name, 0) or 0
                if gun_value > 0:
                    result.append({
                        "bayi_kodu": r.get("bayi_kodu", ""),
                        "unvan": r.get("unvan", ""),
                        "dst": r.get("dst", ""),
                        "dsm": r.get("dsm", ""),
                        "tip": r.get("tip", ""),
                        "sinif": r.get("sinif", ""),
                        "musteri_bakiyesi": r.get("musteri_bakiyesi", 0),
                        "gun_deger": gun_value
                    })
        
        result.sort(key=lambda x: x.get("gun_deger", 0), reverse=True)
        return result
    except Exception as e:
        logger.error(f"Error getting DSM cari bayiler: {e}")
        return []

# Ekip Raporu Verileri - Aylar listesi
@api_router.get("/ekip-raporu/aylar")
async def get_ekip_raporu_aylar():
    try:
        pipeline = [
            {"$group": {"_id": "$ay"}},
            {"$sort": {"_id": 1}}
        ]
        aylar = await db.ekip_raporu.aggregate(pipeline).to_list(20)
        ay_sirasi = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN", 
                     "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]
        result = [a["_id"] for a in aylar if a["_id"]]
        result.sort(key=lambda x: ay_sirasi.index(x) if x in ay_sirasi else 99)
        return result
    except Exception as e:
        logger.error(f"Error getting ekip raporu aylar: {e}")
        return []

# Ekip Raporu Verileri - Ay bazlı günler
@api_router.get("/ekip-raporu/{ay}")
async def get_ekip_raporu_ay(ay: str):
    try:
        records = await db.ekip_raporu.find({"ay": ay.upper()}).to_list(50)
        # Sort by date
        records.sort(key=lambda x: x.get("tarih", 0))
        # Convert ObjectId to string
        for r in records:
            r["_id"] = str(r["_id"])
        return records
    except Exception as e:
        logger.error(f"Error getting ekip raporu for {ay}: {e}")
        return []

# Ekip Raporu Verileri - Yıl toplamları
@api_router.get("/ekip-raporu-toplam")
async def get_ekip_raporu_toplam():
    try:
        toplam = await db.ekip_raporu_toplam.find_one()
        if toplam:
            toplam["_id"] = str(toplam["_id"])
            return toplam
        return {"yil_toplam_karton": {}, "yil_toplam_kasa": {}}
    except Exception as e:
        logger.error(f"Error getting ekip raporu toplam: {e}")
        return {"yil_toplam_karton": {}, "yil_toplam_kasa": {}}

# Loyalty Bayiler
@api_router.get("/loyalty-bayiler")
async def get_loyalty_bayiler():
    try:
        records = await db.loyalty_bayiler.find().to_list(1000)
        for r in records:
            r["_id"] = str(r["_id"])
        return records
    except Exception as e:
        logger.error(f"Error getting loyalty bayiler: {e}")
        return []

# Loyalty Bayi Sayısı
@api_router.get("/loyalty-bayi-sayisi")
async def get_loyalty_bayi_sayisi():
    try:
        count = await db.loyalty_bayiler.count_documents({})
        return {"count": count}
    except Exception as e:
        logger.error(f"Error getting loyalty bayi count: {e}")
        return {"count": 0}

# Bayi Hedef (Aylık marka hedefleri)
@api_router.get("/bayi-hedef/{bayi_kodu}")
async def get_bayi_hedef(bayi_kodu: str):
    try:
        hedef = await db.bayi_hedef.find_one({"bayi_kodu": bayi_kodu})
        if hedef:
            hedef["_id"] = str(hedef["_id"])
            return hedef
        return None
    except Exception as e:
        logger.error(f"Error getting bayi hedef: {e}")
        return None

# Carili Kanal Toplamları
@api_router.get("/carili-kanal-toplamlari")
async def get_carili_kanal_toplamlari():
    try:
        toplam = await db.carili_kanal_toplamlari.find_one()
        if toplam:
            toplam["_id"] = str(toplam["_id"]) if "_id" in toplam else None
            return toplam
        return {}
    except Exception as e:
        logger.error(f"Error getting carili kanal toplamlari: {e}")
        return {}

# Son Güncelleme Zamanı
@api_router.get("/son-guncelleme")
async def get_son_guncelleme():
    try:
        info = await db.system_info.find_one({"type": "excel_upload"})
        if info:
            return {"son_guncelleme": info.get("son_guncelleme", "")}
        return {"son_guncelleme": ""}
    except Exception as e:
        logger.error(f"Error getting son guncelleme: {e}")
        return {"son_guncelleme": ""}

# Kanal Bazlı Müşteri Listesi
@api_router.get("/kanal-musterileri/{kanal}")
async def get_kanal_musterileri(kanal: str):
    try:
        # Stand raporundan kanal bazlı müşteri listesi
        records = await db.stand_raporu.find().to_list(5000)
        result = []
        for r in records:
            # Kanal bilgisi için konya_gun'dan da bakabiliriz
            bayi_kodu = r.get("bayi_kodu", "")
            konya_data = await db.konya_gun.find_one({"bayi_kodu": bayi_kodu})
            
            if konya_data:
                bayi_kanal = konya_data.get("kanal", "").upper().strip() if konya_data.get("kanal") else ""
                if kanal.upper() in bayi_kanal or bayi_kanal in kanal.upper():
                    r["_id"] = str(r["_id"])
                    r["kanal"] = bayi_kanal
                    r["musteri_bakiyesi"] = konya_data.get("musteri_bakiyesi", 0)
                    result.append(r)
        
        return result
    except Exception as e:
        logger.error(f"Error getting kanal musterileri: {e}")
        return []

# İlçe Bazlı Veriler (Harita için)
@api_router.get("/ilce-verileri")
async def get_ilce_verileri():
    try:
        # Stand raporundan ilçe bilgisini al
        records = await db.stand_raporu.find().to_list(5000)
        ilce_data = {}
        
        for r in records:
            ilce = r.get("ilce", "")  # L sütunu
            if not ilce:
                continue
                
            if ilce not in ilce_data:
                ilce_data[ilce] = {
                    "ilce": ilce,
                    "bayi_sayisi": 0,
                    "aktif_bayi": 0,
                    "pasif_bayi": 0,
                    "bayiler": []
                }
            
            ilce_data[ilce]["bayi_sayisi"] += 1
            bayi_durumu = r.get("bayi_durumu", "")
            if bayi_durumu == "AKTİF":
                ilce_data[ilce]["aktif_bayi"] += 1
            else:
                ilce_data[ilce]["pasif_bayi"] += 1
            
            ilce_data[ilce]["bayiler"].append({
                "bayi_kodu": r.get("bayi_kodu", ""),
                "bayi_unvani": r.get("bayi_unvani", ""),
                "bayi_durumu": bayi_durumu
            })
        
        return list(ilce_data.values())
    except Exception as e:
        logger.error(f"Error getting ilce verileri: {e}")
        return []

# Stand Raporu (Kanal Kırılım için)
@api_router.get("/stand-raporu")
async def get_stand_raporu():
    try:
        records = await db.stand_raporu.find().to_list(5000)
        for r in records:
            r["_id"] = str(r["_id"])
        return records
    except Exception as e:
        logger.error(f"Error getting stand raporu: {e}")
        return []

# Kanal Müşterileri - Tip bazlı filtreleme
@api_router.get("/kanal-musterileri/{kanal}")
async def get_kanal_musterileri(kanal: str):
    try:
        # Kanal tipine göre filtreleme
        # 12YZNC = Yerel Zincir
        # 08ASK = Askeriye
        # 11CZV = Cezaevi
        # Geri kalan = Geleneksel (Piyasa, Benzinlik)
        
        query = {}
        kanal_lower = kanal.lower()
        
        if kanal_lower == "yerel-zincir":
            # 12YZNC - tip kodunda "12" başlayanlar
            query = {"tip": {"$regex": "^12", "$options": "i"}}
        elif kanal_lower == "askeriye":
            # 08ASK - tip kodunda "08" başlayanlar
            query = {"tip": {"$regex": "^08", "$options": "i"}}
        elif kanal_lower == "cezaevi":
            # 11CZV - tip kodunda "11" başlayanlar
            query = {"tip": {"$regex": "^11", "$options": "i"}}
        elif kanal_lower == "benzinlik":
            # 07BEN - tip kodunda "07" başlayanlar
            query = {"tip": {"$regex": "^07", "$options": "i"}}
        elif kanal_lower == "piyasa":
            # Piyasa = 01BAK, 02MAR, 03BFE, 04KYE, 05TEK
            query = {"tip": {"$regex": "^(01|02|03|04|05)", "$options": "i"}}
        elif kanal_lower == "geleneksel":
            # Geleneksel = 14TUT, 15TUS
            query = {"tip": {"$regex": "^(14|15)", "$options": "i"}}
        else:
            # Spesifik kod
            query = {"tip": {"$regex": f"^{kanal}", "$options": "i"}}
        
        records = await db.stand_raporu.find(query).to_list(5000)
        for r in records:
            r["_id"] = str(r["_id"])
        return records
    except Exception as e:
        logger.error(f"Error getting kanal musterileri: {e}")
        return []

# Stil Ay Satış
@api_router.get("/stil-ay-satis")
async def get_stil_ay_satis():
    try:
        records = await db.stil_ay_satis.find().to_list(20)
        ay_sirasi = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN", 
                     "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK", "YIL TOPLAM"]
        records.sort(key=lambda x: ay_sirasi.index(x.get("ay", "")) if x.get("ay", "") in ay_sirasi else 99)
        for r in records:
            r["_id"] = str(r["_id"])
        return records
    except Exception as e:
        logger.error(f"Error getting stil ay satis: {e}")
        return []

# Personel Data
@api_router.get("/personel-data")
async def get_personel_data(isim: str = Query(default="", description="İsim filtresi")):
    try:
        query = {}
        if isim:
            isim_ascii = turkish_to_ascii(isim)
            # Get all and filter manually for Turkish-insensitive search
            all_records = await db.personel_data.find().to_list(100)
            result = []
            for r in all_records:
                adi = r.get("adi", "") or ""
                if isim_ascii.lower() in turkish_to_ascii(adi).lower():
                    r["_id"] = str(r["_id"])
                    result.append(r)
            return result
        else:
            records = await db.personel_data.find().to_list(100)
            for r in records:
                r["_id"] = str(r["_id"])
            return records
    except Exception as e:
        logger.error(f"Error getting personel data: {e}")
        return []

# RUT Günleri listesi (DST bazlı)
@api_router.get("/rut/gunler")
async def get_rut_gunler(dst_name: str = Query(default="", description="DST adı filtresi")):
    try:
        pipeline = []
        
        if dst_name:
            dst_ascii = turkish_to_ascii(dst_name)
            # Get all records and filter manually
            all_records = await db.rut_data.find().to_list(5000)
            gunler_set = set()
            for r in all_records:
                db_dst = r.get("dst_name", "") or ""
                if turkish_to_ascii(db_dst).lower() == dst_ascii.lower():
                    gun = r.get("gun", "")
                    if gun:
                        gunler_set.add(gun)
            
            # Sort days in correct order
            gun_sirasi = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
            result = sorted(list(gunler_set), key=lambda x: gun_sirasi.index(x) if x in gun_sirasi else 99)
            return result
        else:
            # Get all unique days
            pipeline = [
                {"$group": {"_id": "$gun"}},
                {"$match": {"_id": {"$ne": None, "$ne": ""}}}
            ]
            gunler = await db.rut_data.aggregate(pipeline).to_list(10)
            gun_sirasi = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
            result = [g["_id"] for g in gunler if g["_id"]]
            result.sort(key=lambda x: gun_sirasi.index(x) if x in gun_sirasi else 99)
            return result
    except Exception as e:
        logger.error(f"Error getting rut gunler: {e}")
        return []

# RUT DST listesi
@api_router.get("/rut/dst-list")
async def get_rut_dst_list():
    try:
        pipeline = [
            {"$group": {"_id": "$dst_name"}},
            {"$match": {"_id": {"$ne": None, "$ne": ""}}},
            {"$sort": {"_id": 1}}
        ]
        dstler = await db.rut_data.aggregate(pipeline).to_list(50)
        return [d["_id"] for d in dstler if d["_id"]]
    except Exception as e:
        logger.error(f"Error getting rut dst list: {e}")
        return []

# RUT verileri (DST + gün bazlı)
@api_router.get("/rut")
async def get_rut_data(dst_name: str = Query(..., description="DST adı"), gun: str = Query(..., description="Gün")):
    try:
        dst_ascii = turkish_to_ascii(dst_name)
        gun_lower = gun.lower()
        
        # Get all records and filter
        all_records = await db.rut_data.find().to_list(5000)
        result = []
        
        for r in all_records:
            db_dst = r.get("dst_name", "") or ""
            db_gun = r.get("gun", "") or ""
            
            if turkish_to_ascii(db_dst).lower() == dst_ascii.lower() and db_gun.lower() == gun_lower:
                r["_id"] = str(r["_id"])
                result.append(r)
        
        # Sort by ziyaret_sira
        result.sort(key=lambda x: x.get("ziyaret_sira", 0))
        return result
    except Exception as e:
        logger.error(f"Error getting rut data: {e}")
        return []

# RUT Talep Gönder
@api_router.post("/rut/talep")
async def send_rut_talep(request: RutTalepRequest):
    try:
        import io
        import json
        from datetime import datetime
        
        # Create talep record
        talep = {
            "dst_name": request.dst_name,
            "gun": request.gun,
            "yeni_sira": request.yeni_sira,
            "tarih": datetime.now(),
            "durum": "beklemede",  # beklemede, onaylandi, reddedildi
        }
        
        result = await db.rut_talepler.insert_one(talep)
        
        logger.info(f"RUT talebi oluşturuldu: {request.dst_name} - {request.gun}")
        
        return {
            "success": True,
            "message": f"{request.dst_name} için {request.gun} günü rut değişiklik talebi gönderildi",
            "talep_id": str(result.inserted_id)
        }
    except Exception as e:
        logger.error(f"Error sending rut talep: {e}")
        return {"success": False, "message": str(e)}

# RUT Talepleri Listesi (Admin için)
@api_router.get("/rut/talepler")
async def get_rut_talepler():
    try:
        talepler = await db.rut_talepler.find().sort("tarih", -1).to_list(100)
        for t in talepler:
            t["_id"] = str(t["_id"])
            if t.get("tarih"):
                t["tarih"] = t["tarih"].isoformat()
        return talepler
    except Exception as e:
        logger.error(f"Error getting rut talepler: {e}")
        return []

# RUT Talep sayısı (Bekleyenler)
@api_router.get("/rut/talep-sayisi")
async def get_rut_talep_sayisi():
    try:
        count = await db.rut_talepler.count_documents({"durum": "beklemede"})
        return {"count": count}
    except Exception as e:
        logger.error(f"Error getting talep sayisi: {e}")
        return {"count": 0}

# RUT Talep Excel İndir
@api_router.get("/rut/talep/{talep_id}/excel")
async def download_rut_talep_excel(talep_id: str):
    try:
        from bson import ObjectId
        from fastapi.responses import StreamingResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io
        import re
        
        talep = await db.rut_talepler.find_one({"_id": ObjectId(talep_id)})
        if not talep:
            raise HTTPException(status_code=404, detail="Talep bulunamadı")
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "RUT Sıralaması"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header row
        headers = ["Sıra", "Müşteri Kodu", "Müşteri Ünvanı", "Durum", "Grup"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, item in enumerate(talep.get("yeni_sira", []), 2):
            ws.cell(row=row_idx, column=1, value=item.get("ziyaret_sira", "")).border = thin_border
            ws.cell(row=row_idx, column=2, value=item.get("musteri_kod", "")).border = thin_border
            ws.cell(row=row_idx, column=3, value=item.get("musteri_unvan", "")).border = thin_border
            ws.cell(row=row_idx, column=4, value=item.get("musteri_durum", "")).border = thin_border
            ws.cell(row=row_idx, column=5, value=item.get("musteri_grup", "")).border = thin_border
        
        # Column widths
        ws.column_dimensions['A'].width = 8   # Sıra
        ws.column_dimensions['B'].width = 15  # Müşteri Kodu
        ws.column_dimensions['C'].width = 40  # Müşteri Ünvanı
        ws.column_dimensions['D'].width = 12  # Durum
        ws.column_dimensions['E'].width = 15  # Grup
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # ASCII-safe filename for header
        def safe_filename(s):
            tr_map = {'ı': 'i', 'İ': 'I', 'ş': 's', 'Ş': 'S', 'ğ': 'g', 'Ğ': 'G', 
                      'ü': 'u', 'Ü': 'U', 'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C'}
            for tr_char, en_char in tr_map.items():
                s = s.replace(tr_char, en_char)
            s = re.sub(r'[^\w\-]', '_', s)
            return s
        
        dst_name = safe_filename(talep.get("dst_name", "DST"))
        gun = safe_filename(talep.get("gun", "Gun"))
        filename = f"RUT_{dst_name}_{gun}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
    except Exception as e:
        logger.error(f"Error downloading talep excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# RUT Talep Durumu Güncelle (Admin)
@api_router.put("/rut/talep/{talep_id}")
async def update_rut_talep(talep_id: str, durum: str = Query(..., description="Yeni durum: onaylandi, reddedildi")):
    try:
        from bson import ObjectId
        
        result = await db.rut_talepler.update_one(
            {"_id": ObjectId(talep_id)},
            {"$set": {"durum": durum}}
        )
        
        if result.modified_count > 0:
            return {"success": True, "message": f"Talep durumu '{durum}' olarak güncellendi"}
        return {"success": False, "message": "Talep bulunamadı"}
    except Exception as e:
        logger.error(f"Error updating talep: {e}")
        return {"success": False, "message": str(e)}

# Bayi search
@api_router.get("/bayiler", response_model=List[BayiSummary])
async def search_bayiler(q: str = Query(default="", description="Search query")):
    try:
        query = {}
        if q:
            # Convert search term to ASCII for Turkish-insensitive search
            search_ascii = turkish_to_ascii(q)
            query = {
                "$or": [
                    {"bayi_kodu_ascii": {"$regex": search_ascii, "$options": "i"}},
                    {"bayi_unvani_ascii": {"$regex": search_ascii, "$options": "i"}}
                ]
            }
        
        # Sort by kapsam_durumu: Aktif first, then Pasif, then İptal
        pipeline = [
            {"$match": query} if query else {"$match": {}},
            {"$addFields": {
                "sort_order": {
                    "$switch": {
                        "branches": [
                            {"case": {"$eq": ["$kapsam_durumu", "Aktif"]}, "then": 1},
                            {"case": {"$eq": ["$kapsam_durumu", "Pasif"]}, "then": 2},
                            {"case": {"$eq": ["$kapsam_durumu", "İptal"]}, "then": 3}
                        ],
                        "default": 4
                    }
                }
            }},
            {"$sort": {"sort_order": 1, "bayi_unvani": 1}},
            {"$limit": 100}
        ]
        
        bayiler = await db.bayiler.aggregate(pipeline).to_list(100)
        return [BayiSummary(
            bayi_kodu=str(b.get("bayi_kodu", "")),
            bayi_unvani=b.get("bayi_unvani", ""),
            kapsam_durumu=b.get("kapsam_durumu"),
            tip=b.get("tip"),
            sinif=b.get("panaroma_sinif"),
            dst=b.get("dst"),
            tte=b.get("tte")
        ) for b in bayiler]
    except Exception as e:
        logger.error(f"Error searching bayiler: {e}")
        return []

# Bayi detail
@api_router.get("/bayiler/{bayi_kodu}", response_model=BayiDetail)
async def get_bayi_detail(bayi_kodu: str):
    try:
        bayi = await db.bayiler.find_one({"bayi_kodu": bayi_kodu})
        if not bayi:
            # Check if any data exists in the collection
            count = await db.bayiler.count_documents({})
            if count == 0:
                raise HTTPException(status_code=404, detail="Veri yüklenmemiş. Lütfen Excel dosyasını yükleyin.")
            raise HTTPException(status_code=404, detail="Bayi bulunamadı")
        
        # Get borç durumu from konya_gun collection
        borc = await db.konya_gun.find_one({"bayi_kodu": bayi_kodu})
        borc_durumu = "Borcu yoktur"
        if borc and borc.get("musteri_bakiyesi"):
            bakiye = safe_float(borc.get("musteri_bakiyesi"))
            if bakiye > 0:
                borc_durumu = f"{bakiye:,.1f} TL"
        
        # Calculate development percentage
        toplam_2024 = safe_float(bayi.get("toplam_satis_2024"))
        toplam_2025 = safe_float(bayi.get("toplam_satis_2025"))
        gelisim = 0.0
        if toplam_2024 > 0:
            gelisim = ((toplam_2025 - toplam_2024) / toplam_2024) * 100
        
        # Get ziyaret günleri from stand_raporu collection
        stand = await db.stand_raporu.find_one({"bayi_kodu": bayi_kodu})
        ziyaret_gunleri = stand.get("ziyaret_gunleri", []) if stand else []
        
        return BayiDetail(
            bayi_kodu=str(bayi.get("bayi_kodu", "")),
            bayi_unvani=bayi.get("bayi_unvani", ""),
            dst=bayi.get("dst"),
            tte=bayi.get("tte"),
            dsm=bayi.get("dsm"),
            tip=bayi.get("tip"),
            panaroma_sinif=bayi.get("panaroma_sinif"),
            satisa_gore_sinif=bayi.get("satisa_gore_sinif"),
            kapsam_durumu=bayi.get("kapsam_durumu"),
            jti_stant=bayi.get("jti_stant"),
            jti_stant_adet=safe_float(bayi.get("jti_stant_adet")),
            camel_myo_stant=bayi.get("camel_myo_stant"),
            camel_myo_adet=safe_float(bayi.get("camel_myo_adet")),
            pmi_stant=bayi.get("pmi_stant"),
            pmi_adet=safe_float(bayi.get("pmi_adet")),
            bat_stant=bayi.get("bat_stant"),
            bat_adet=safe_float(bayi.get("bat_adet")),
            loyalty_plan_2025=safe_float(bayi.get("loyalty_plan_2025")),
            odenen_2025=safe_float(bayi.get("odenen_2025")),
            ocak_2025=safe_float(bayi.get("ocak_2025")),
            subat_2025=safe_float(bayi.get("subat_2025")),
            mart_2025=safe_float(bayi.get("mart_2025")),
            nisan_2025=safe_float(bayi.get("nisan_2025")),
            mayis_2025=safe_float(bayi.get("mayis_2025")),
            haziran_2025=safe_float(bayi.get("haziran_2025")),
            temmuz_2025=safe_float(bayi.get("temmuz_2025")),
            agustos_2025=safe_float(bayi.get("agustos_2025")),
            eylul_2025=safe_float(bayi.get("eylul_2025")),
            ekim_2025=safe_float(bayi.get("ekim_2025")),
            kasim_2025=safe_float(bayi.get("kasim_2025")),
            aralik_2025=safe_float(bayi.get("aralik_2025")),
            toplam_satis_2025=safe_float(bayi.get("toplam_satis_2025")),
            ortalama_2025=safe_float(bayi.get("ortalama_2025")),
            ocak_2024=safe_float(bayi.get("ocak_2024")),
            subat_2024=safe_float(bayi.get("subat_2024")),
            mart_2024=safe_float(bayi.get("mart_2024")),
            nisan_2024=safe_float(bayi.get("nisan_2024")),
            mayis_2024=safe_float(bayi.get("mayis_2024")),
            haziran_2024=safe_float(bayi.get("haziran_2024")),
            temmuz_2024=safe_float(bayi.get("temmuz_2024")),
            agustos_2024=safe_float(bayi.get("agustos_2024")),
            eylul_2024=safe_float(bayi.get("eylul_2024")),
            ekim_2024=safe_float(bayi.get("ekim_2024")),
            kasim_2024=safe_float(bayi.get("kasim_2024")),
            aralik_2024=safe_float(bayi.get("aralik_2024")),
            toplam_satis_2024=safe_float(bayi.get("toplam_satis_2024")),
            ortalama_2024=safe_float(bayi.get("ortalama_2024")),
            ocak_2026=safe_float(bayi.get("ocak_2026")),
            subat_2026=safe_float(bayi.get("subat_2026")),
            mart_2026=safe_float(bayi.get("mart_2026")),
            nisan_2026=safe_float(bayi.get("nisan_2026")),
            mayis_2026=safe_float(bayi.get("mayis_2026")),
            haziran_2026=safe_float(bayi.get("haziran_2026")),
            temmuz_2026=safe_float(bayi.get("temmuz_2026")),
            agustos_2026=safe_float(bayi.get("agustos_2026")),
            eylul_2026=safe_float(bayi.get("eylul_2026")),
            ekim_2026=safe_float(bayi.get("ekim_2026")),
            kasim_2026=safe_float(bayi.get("kasim_2026")),
            aralik_2026=safe_float(bayi.get("aralik_2026")),
            toplam_2026=safe_float(bayi.get("toplam_2026")),
            ortalama_2026=safe_float(bayi.get("ortalama_2026")),
            gelisim_yuzdesi=gelisim,
            borc_durumu=borc_durumu,
            ziyaret_gunleri=ziyaret_gunleri
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting bayi detail: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Faturalar for a bayi
@api_router.get("/bayiler/{bayi_kodu}/faturalar", response_model=List[Fatura])
async def get_bayi_faturalar(bayi_kodu: str):
    try:
        # Sort by date descending (newest first)
        faturalar = await db.faturalar.find({"bayi_kodu": bayi_kodu}).sort("tarih_sort", -1).to_list(1000)
        return [Fatura(
            matbu_no=f.get("matbu_no", ""),
            tarih=f.get("tarih", ""),
            net_tutar=safe_float(f.get("net_tutar")),
            bayi_kodu=f.get("bayi_kodu", "")
        ) for f in faturalar]
    except Exception as e:
        logger.error(f"Error getting faturalar: {e}")
        return []

# Fatura detail with products
@api_router.get("/faturalar/{matbu_no}", response_model=FaturaDetay)
async def get_fatura_detail(matbu_no: str):
    try:
        detaylar = await db.belge_detay.find({"matbu_no": matbu_no}).to_list(1000)
        urunler = []
        toplam_miktar = 0.0
        toplam_tutar = 0.0
        for d in detaylar:
            miktar = safe_float(d.get("miktar"))
            birim_fiyat = safe_float(d.get("birim_fiyat"))
            tutar = miktar * birim_fiyat
            toplam_miktar += miktar
            toplam_tutar += tutar
            urunler.append({
                "urun_adi": d.get("urun", ""),
                "miktar": miktar,
                "birim_fiyat": birim_fiyat,
                "tutar": tutar
            })
        return FaturaDetay(matbu_no=matbu_no, urunler=urunler, toplam_miktar=toplam_miktar, toplam_tutar=toplam_tutar)
    except Exception as e:
        logger.error(f"Error getting fatura detail: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Tahsilatlar for a bayi
@api_router.get("/bayiler/{bayi_kodu}/tahsilatlar", response_model=List[Tahsilat])
async def get_bayi_tahsilatlar(bayi_kodu: str):
    try:
        # Sort by date descending (newest first)
        tahsilatlar = await db.tahsilatlar.find({"bayi_kodu": bayi_kodu}).sort("tarih_sort", -1).to_list(1000)
        return [Tahsilat(
            tahsilat_turu=t.get("tahsilat_turu", ""),
            islem_tarihi=t.get("islem_tarihi", ""),
            tutar=safe_float(t.get("tutar")),
            bayi_kodu=t.get("bayi_kodu", "")
        ) for t in tahsilatlar]
    except Exception as e:
        logger.error(f"Error getting tahsilatlar: {e}")
        return []

# Excel upload endpoint
@api_router.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    try:
        logger.info(f"Receiving file: {file.filename}")
        
        # Save uploaded file to temp
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsb') as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        logger.info(f"File saved to {tmp_path}, size: {len(content)} bytes")
        
        # Process the Excel file
        await process_excel(tmp_path)
        
        # Clean up
        os.unlink(tmp_path)
        
        return {"success": True, "message": "Excel dosyası başarıyla yüklendi ve işlendi"}
    except Exception as e:
        logger.error(f"Upload error: {e}")
        raise HTTPException(status_code=500, detail=f"Yükleme hatası: {str(e)}")

async def process_excel(file_path: str):
    """Process the Excel file and populate MongoDB collections"""
    logger.info("Starting Excel processing...")
    
    # Clear existing data
    await db.bayiler.delete_many({})
    await db.faturalar.delete_many({})
    await db.belge_detay.delete_many({})
    await db.tahsilatlar.delete_many({})
    await db.konya_gun.delete_many({})
    await db.stand_raporu.delete_many({})
    await db.dst_data.delete_many({})
    await db.dsm_teams.delete_many({})
    await db.tte_data.delete_many({})
    
    with pyxlsb.open_workbook(file_path) as wb:
        # Process AÜ BAYİ LİST
        logger.info("Processing AÜ BAYİ LİST...")
        with wb.get_sheet('AÜ BAYİ LİST') as sheet:
            rows = list(sheet.rows())
            bayiler_data = []
            
            for row in rows[2:]:  # Start from row 3 (index 2)
                cells = [cell.v for cell in row]
                if len(cells) > 0 and cells[0]:
                    bayi_kodu = str(int(cells[0])) if isinstance(cells[0], float) else str(cells[0])
                    bayi_unvani = safe_str(cells[1]) if len(cells) > 1 else ""
                    
                    bayi = {
                        "bayi_kodu": bayi_kodu,
                        "bayi_kodu_ascii": turkish_to_ascii(bayi_kodu),
                        "bayi_unvani": bayi_unvani,
                        "bayi_unvani_ascii": turkish_to_ascii(bayi_unvani) if bayi_unvani else "",
                        "dst": safe_str(cells[2]) if len(cells) > 2 else None,
                        "tte": safe_str(cells[3]) if len(cells) > 3 else None,
                        "dsm": safe_str(cells[4]) if len(cells) > 4 else None,
                        "tip": safe_str(cells[5]) if len(cells) > 5 else None,
                        "panaroma_sinif": safe_str(cells[6]) if len(cells) > 6 else None,
                        "satisa_gore_sinif": safe_str(cells[7]) if len(cells) > 7 else None,
                        "kapsam_durumu": safe_str(cells[9]) if len(cells) > 9 else None,
                        "jti_stant": safe_str(cells[10]) if len(cells) > 10 else None,
                        "jti_stant_adet": safe_float(cells[11]) if len(cells) > 11 else 0,
                        "camel_myo_stant": safe_str(cells[12]) if len(cells) > 12 else None,
                        "camel_myo_adet": safe_float(cells[13]) if len(cells) > 13 else 0,
                        "pmi_stant": safe_str(cells[14]) if len(cells) > 14 else None,
                        "pmi_adet": safe_float(cells[15]) if len(cells) > 15 else 0,
                        "bat_stant": safe_str(cells[16]) if len(cells) > 16 else None,
                        "bat_adet": safe_float(cells[17]) if len(cells) > 17 else 0,
                        "loyalty_plan_2025": safe_float(cells[18]) if len(cells) > 18 else 0,
                        "odenen_2025": safe_float(cells[19]) if len(cells) > 19 else 0,
                        # 2025 Monthly
                        "ocak_2025": safe_float(cells[33]) if len(cells) > 33 else 0,
                        "subat_2025": safe_float(cells[34]) if len(cells) > 34 else 0,
                        "mart_2025": safe_float(cells[35]) if len(cells) > 35 else 0,
                        "nisan_2025": safe_float(cells[36]) if len(cells) > 36 else 0,
                        "mayis_2025": safe_float(cells[37]) if len(cells) > 37 else 0,
                        "haziran_2025": safe_float(cells[38]) if len(cells) > 38 else 0,
                        "temmuz_2025": safe_float(cells[39]) if len(cells) > 39 else 0,
                        "agustos_2025": safe_float(cells[40]) if len(cells) > 40 else 0,
                        "eylul_2025": safe_float(cells[41]) if len(cells) > 41 else 0,
                        "ekim_2025": safe_float(cells[42]) if len(cells) > 42 else 0,
                        "kasim_2025": safe_float(cells[43]) if len(cells) > 43 else 0,
                        "aralik_2025": safe_float(cells[44]) if len(cells) > 44 else 0,
                        "toplam_satis_2025": safe_float(cells[45]) if len(cells) > 45 else 0,
                        "ortalama_2025": safe_float(cells[48]) if len(cells) > 48 else 0,
                        # 2024 Monthly
                        "ocak_2024": safe_float(cells[71]) if len(cells) > 71 else 0,
                        "subat_2024": safe_float(cells[72]) if len(cells) > 72 else 0,
                        "mart_2024": safe_float(cells[73]) if len(cells) > 73 else 0,
                        "nisan_2024": safe_float(cells[74]) if len(cells) > 74 else 0,
                        "mayis_2024": safe_float(cells[75]) if len(cells) > 75 else 0,
                        "haziran_2024": safe_float(cells[76]) if len(cells) > 76 else 0,
                        "temmuz_2024": safe_float(cells[77]) if len(cells) > 77 else 0,
                        "agustos_2024": safe_float(cells[78]) if len(cells) > 78 else 0,
                        "eylul_2024": safe_float(cells[79]) if len(cells) > 79 else 0,
                        "ekim_2024": safe_float(cells[80]) if len(cells) > 80 else 0,
                        "kasim_2024": safe_float(cells[81]) if len(cells) > 81 else 0,
                        "aralik_2024": safe_float(cells[82]) if len(cells) > 82 else 0,
                        "toplam_satis_2024": safe_float(cells[83]) if len(cells) > 83 else 0,
                        "ortalama_2024": safe_float(cells[84]) if len(cells) > 84 else 0,
                        # 2026 Monthly
                        "ocak_2026": safe_float(cells[85]) if len(cells) > 85 else 0,
                        "subat_2026": safe_float(cells[86]) if len(cells) > 86 else 0,
                        "mart_2026": safe_float(cells[87]) if len(cells) > 87 else 0,
                        "nisan_2026": safe_float(cells[88]) if len(cells) > 88 else 0,
                        "mayis_2026": safe_float(cells[89]) if len(cells) > 89 else 0,
                        "haziran_2026": safe_float(cells[90]) if len(cells) > 90 else 0,
                        "temmuz_2026": safe_float(cells[91]) if len(cells) > 91 else 0,
                        "agustos_2026": safe_float(cells[92]) if len(cells) > 92 else 0,
                        "eylul_2026": safe_float(cells[93]) if len(cells) > 93 else 0,
                        "ekim_2026": safe_float(cells[94]) if len(cells) > 94 else 0,
                        "kasim_2026": safe_float(cells[95]) if len(cells) > 95 else 0,
                        "aralik_2026": safe_float(cells[96]) if len(cells) > 96 else 0,
                        "toplam_2026": safe_float(cells[97]) if len(cells) > 97 else 0,
                        "ortalama_2026": safe_float(cells[98]) if len(cells) > 98 else 0,
                    }
                    bayiler_data.append(bayi)
            
            if bayiler_data:
                await db.bayiler.insert_many(bayiler_data)
                logger.info(f"Inserted {len(bayiler_data)} bayiler")
        
        # Process Fatura
        logger.info("Processing Fatura...")
        with wb.get_sheet('Fatura') as sheet:
            rows = list(sheet.rows())
            faturalar_data = []
            
            for row in rows[1:]:  # Start from row 2
                cells = [cell.v for cell in row]
                if len(cells) > 13 and cells[0]:
                    bayi_kodu = str(cells[0]).strip() if cells[0] else ""
                    tarih = excel_date_to_str(cells[3])
                    
                    # Parse date for sorting
                    tarih_sort = None
                    if cells[3]:
                        try:
                            if isinstance(cells[3], (int, float)):
                                tarih_sort = int(cells[3])
                            else:
                                # Try to parse date string
                                parts = str(cells[3]).replace('.', '/').replace('-', '/').split('/')
                                if len(parts) == 3:
                                    tarih_sort = int(parts[2]) * 10000 + int(parts[1]) * 100 + int(parts[0])
                        except:
                            tarih_sort = 0
                    
                    fatura = {
                        "bayi_kodu": bayi_kodu,
                        "tarih": tarih,
                        "tarih_sort": tarih_sort or 0,
                        "matbu_no": safe_str(cells[5]) if len(cells) > 5 else "",
                        "net_tutar": safe_float(cells[13]) if len(cells) > 13 else 0
                    }
                    faturalar_data.append(fatura)
            
            if faturalar_data:
                await db.faturalar.insert_many(faturalar_data)
                logger.info(f"Inserted {len(faturalar_data)} faturalar")
        
        # Process Belge Detay
        logger.info("Processing Belge detay...")
        with wb.get_sheet('Belge detay') as sheet:
            rows = list(sheet.rows())
            detay_data = []
            
            for row in rows[1:]:  # Start from row 2
                cells = [cell.v for cell in row]
                if len(cells) > 7 and cells[0]:
                    detay = {
                        "matbu_no": safe_str(cells[0]),
                        "urun": safe_str(cells[6]) if len(cells) > 6 else "",
                        "miktar": safe_float(cells[7]) if len(cells) > 7 else 0,
                        "birim_fiyat": safe_float(cells[8]) if len(cells) > 8 else 0
                    }
                    detay_data.append(detay)
            
            if detay_data:
                await db.belge_detay.insert_many(detay_data)
                logger.info(f"Inserted {len(detay_data)} belge detay")
        
        # Process Tahsilat
        logger.info("Processing tahsilat...")
        with wb.get_sheet('tahsilat') as sheet:
            rows = list(sheet.rows())
            tahsilat_data = []
            
            for row in rows[1:]:  # Start from row 2
                cells = [cell.v for cell in row]
                if len(cells) > 8 and cells[2]:
                    bayi_kodu = str(cells[2]).strip() if cells[2] else ""
                    islem_tarihi = safe_str(cells[5])
                    
                    # Parse date for sorting
                    tarih_sort = 0
                    if islem_tarihi:
                        try:
                            parts = islem_tarihi.replace('.', '/').replace('-', '/').split('/')
                            if len(parts) == 3:
                                # Format: DD/MM/YYYY
                                tarih_sort = int(parts[2]) * 10000 + int(parts[1]) * 100 + int(parts[0])
                        except:
                            tarih_sort = 0
                    
                    tahsilat = {
                        "bayi_kodu": bayi_kodu,
                        "tahsilat_turu": safe_str(cells[1]) if len(cells) > 1 else "",
                        "islem_tarihi": islem_tarihi,
                        "tarih_sort": tarih_sort,
                        "tutar": safe_float(cells[8]) if len(cells) > 8 else 0
                    }
                    tahsilat_data.append(tahsilat)
            
            if tahsilat_data:
                await db.tahsilatlar.insert_many(tahsilat_data)
                logger.info(f"Inserted {len(tahsilat_data)} tahsilatlar")
        
        # Process Konya Gün (for debt info)
        logger.info("Processing KONYA GÜN...")
        with wb.get_sheet('KONYA GÜN') as sheet:
            rows = list(sheet.rows())
            konya_data = []
            
            # Carili kanal toplamları (H2-H6, K2-K6)
            carili_kanal_toplamlari = {}
            if len(rows) > 5:
                for i in range(1, 6):  # rows 2-6 (index 1-5)
                    if i < len(rows):
                        row_cells = [cell.v for cell in rows[i]]
                        if len(row_cells) > 10:
                            kanal_adi = safe_str(row_cells[7]) if len(row_cells) > 7 else ""  # H column
                            kanal_tutari = safe_float(row_cells[10]) if len(row_cells) > 10 else 0  # K column
                            if kanal_adi:
                                carili_kanal_toplamlari[kanal_adi.upper().strip()] = kanal_tutari
            
            # Store carili totals
            if carili_kanal_toplamlari:
                await db.carili_kanal_toplamlari.delete_many({})
                await db.carili_kanal_toplamlari.insert_one(carili_kanal_toplamlari)
                logger.info(f"Saved carili kanal toplamları: {carili_kanal_toplamlari}")
            
            for row in rows[7:]:  # Start from row 8 (index 7) - after header
                cells = [cell.v for cell in row]
                if len(cells) > 10 and cells[0]:
                    bayi_kodu = str(int(cells[0])) if isinstance(cells[0], float) else str(cells[0])
                    konya = {
                        "bayi_kodu": bayi_kodu,
                        "unvan": safe_str(cells[1]) if len(cells) > 1 else "",
                        "dst": safe_str(cells[2]) if len(cells) > 2 else "",
                        "dsm": safe_str(cells[3]) if len(cells) > 3 else "",
                        "tip": safe_str(cells[4]) if len(cells) > 4 else "",
                        "sinif": safe_str(cells[5]) if len(cells) > 5 else "",
                        "kanal": safe_str(cells[7]) if len(cells) > 7 else "",  # H column - kanal bilgisi
                        "musteri_bakiyesi": safe_float(cells[10]) if len(cells) > 10 else 0,
                        "gun_0": safe_float(cells[11]) if len(cells) > 11 else 0,
                        "gun_1": safe_float(cells[12]) if len(cells) > 12 else 0,
                        "gun_2": safe_float(cells[13]) if len(cells) > 13 else 0,
                        "gun_3": safe_float(cells[14]) if len(cells) > 14 else 0,
                        "gun_4": safe_float(cells[15]) if len(cells) > 15 else 0,
                        "gun_5": safe_float(cells[16]) if len(cells) > 16 else 0,
                        "gun_6": safe_float(cells[17]) if len(cells) > 17 else 0,
                        "gun_7": safe_float(cells[18]) if len(cells) > 18 else 0,
                        "gun_8": safe_float(cells[19]) if len(cells) > 19 else 0,
                        "gun_9": safe_float(cells[20]) if len(cells) > 20 else 0,
                        "gun_10": safe_float(cells[21]) if len(cells) > 21 else 0,
                        "gun_11": safe_float(cells[22]) if len(cells) > 22 else 0,
                        "gun_12": safe_float(cells[23]) if len(cells) > 23 else 0,
                        "gun_13": safe_float(cells[24]) if len(cells) > 24 else 0,
                        "gun_14_uzeri": safe_float(cells[25]) if len(cells) > 25 else 0,
                    }
                    konya_data.append(konya)
            
            if konya_data:
                await db.konya_gun.insert_many(konya_data)
                logger.info(f"Inserted {len(konya_data)} konya_gun records")
        
        # Process Stand Raporu (for active/passive counts and visit days)
        logger.info("Processing STAND RAPORU...")
        with wb.get_sheet('STAND RAPORU') as sheet:
            rows = list(sheet.rows())
            stand_data = []
            
            for row in rows[1:]:  # Start from row 2
                cells = [cell.v for cell in row]
                if len(cells) > 72 and cells[5]:
                    bayi_kodu = str(cells[5]).strip() if cells[5] else ""
                    
                    # Bayi Unvanı - column 6 (UNVAN)
                    bayi_unvani = safe_str(cells[6]) if len(cells) > 6 else None
                    
                    # DST - column 59 (BH)
                    dst = safe_str(cells[59]) if len(cells) > 59 else None
                    
                    # TTE - column 60 (BI)
                    tte = safe_str(cells[60]) if len(cells) > 60 else None
                    
                    # TXTKAPSAM - BK sütunu (index 62)
                    txtkapsam = safe_str(cells[62]) if len(cells) > 62 else None
                    
                    # Ziyaret günleri - sütun 66-72
                    ziyaret_gunleri = []
                    gun_isimleri = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
                    for i, gun in enumerate(gun_isimleri):
                        col_idx = 66 + i
                        if col_idx < len(cells) and cells[col_idx] == 1.0:
                            ziyaret_gunleri.append(gun)
                    
                    stand = {
                        "bayi_kodu": bayi_kodu,
                        "bayi_unvani": bayi_unvani,
                        "bayi_durumu": safe_str(cells[12]) if len(cells) > 12 else None,
                        "dst": dst,
                        "tte": tte,
                        "txtkapsam": txtkapsam,
                        "ziyaret_gunleri": ziyaret_gunleri
                    }
                    stand_data.append(stand)
            
            if stand_data:
                await db.stand_raporu.insert_many(stand_data)
                logger.info(f"Inserted {len(stand_data)} stand_raporu records")
        
        # Process DATA (DST verileri)
        logger.info("Processing DATA...")
        with wb.get_sheet('DATA') as sheet:
            rows = list(sheet.rows())
            dst_data_list = []
            
            for row in rows[1:21]:  # 2-21 satırlar (DST verileri, TOPLAM hariç)
                cells = [cell.v for cell in row]
                if len(cells) > 0 and cells[0] and cells[0] not in ['TOPLAM', 'TEAM-I', 'TEAM-II']:
                    dst = safe_str(cells[0])
                    if dst:
                        dst_record = {
                            "dst": dst,
                            "bayi_sayisi": safe_float(cells[1]) if len(cells) > 1 else 0,
                            "aktif_bayi_sayisi": safe_float(cells[2]) if len(cells) > 2 else 0,
                            "pasif_bayi_sayisi": safe_float(cells[3]) if len(cells) > 3 else 0,
                            "aralik_hedef": safe_float(cells[4]) if len(cells) > 4 else 0,
                            "aralik_satis": safe_float(cells[5]) if len(cells) > 5 else 0,
                            "kalan_satis": safe_float(cells[6]) if len(cells) > 6 else 0,
                            "hedef_basari_orani": safe_float(cells[7]) if len(cells) > 7 else 0,
                            "tahsilat_hedef": safe_float(cells[8]) if len(cells) > 8 else 0,
                            "tahsilat_tutari": safe_float(cells[9]) if len(cells) > 9 else 0,
                            "ay_hedef_ziyaret": safe_float(cells[10]) if len(cells) > 10 else 0,
                            "ziyaret_gerceklesen": safe_float(cells[11]) if len(cells) > 11 else 0,
                            "drop_rate": safe_float(cells[12]) if len(cells) > 12 else 0,
                            "basarili_satis": safe_float(cells[13]) if len(cells) > 13 else 0,
                            "basarili_satis_yuzde": safe_float(cells[14]) if len(cells) > 14 else 0,
                            "carili_bayi_sayisi": safe_float(cells[15]) if len(cells) > 15 else 0,
                            "gun_0": safe_float(cells[16]) if len(cells) > 16 else 0,
                            "gun_1": safe_float(cells[17]) if len(cells) > 17 else 0,
                            "gun_2": safe_float(cells[18]) if len(cells) > 18 else 0,
                            "gun_3": safe_float(cells[19]) if len(cells) > 19 else 0,
                            "gun_4": safe_float(cells[20]) if len(cells) > 20 else 0,
                            "gun_5": safe_float(cells[21]) if len(cells) > 21 else 0,
                            "gun_6": safe_float(cells[22]) if len(cells) > 22 else 0,
                            "gun_7": safe_float(cells[23]) if len(cells) > 23 else 0,
                            "gun_8": safe_float(cells[24]) if len(cells) > 24 else 0,
                            "gun_9": safe_float(cells[25]) if len(cells) > 25 else 0,
                            "gun_10": safe_float(cells[26]) if len(cells) > 26 else 0,
                            "gun_11": safe_float(cells[27]) if len(cells) > 27 else 0,
                            "gun_12": safe_float(cells[28]) if len(cells) > 28 else 0,
                            "gun_13": safe_float(cells[29]) if len(cells) > 29 else 0,
                            "gun_14_uzeri": safe_float(cells[30]) if len(cells) > 30 else 0,
                            "cari_toplam": safe_float(cells[31]) if len(cells) > 31 else 0,
                            "loy_verilen_bayi_sayisi": safe_float(cells[32]) if len(cells) > 32 else 0,
                            "loy_bayi_mahsuplasma_tutari": safe_float(cells[33]) if len(cells) > 33 else 0,
                            # SKU Satışları (34-71)
                            "skt_camel_yellow_100": safe_float(cells[34]) if len(cells) > 34 else 0,
                            "camel_brown": safe_float(cells[35]) if len(cells) > 35 else 0,
                            "camel_black": safe_float(cells[36]) if len(cells) > 36 else 0,
                            "camel_white": safe_float(cells[37]) if len(cells) > 37 else 0,
                            "camel_yellow_sp": safe_float(cells[38]) if len(cells) > 38 else 0,
                            "camel_yellow": safe_float(cells[39]) if len(cells) > 39 else 0,
                            "camel_deep_blue_long": safe_float(cells[40]) if len(cells) > 40 else 0,
                            "camel_deep_blue": safe_float(cells[41]) if len(cells) > 41 else 0,
                            "camel_yellow_long": safe_float(cells[42]) if len(cells) > 42 else 0,
                            "camel_slender_blue": safe_float(cells[43]) if len(cells) > 43 else 0,
                            "dp_camel_slender_blueline": safe_float(cells[44]) if len(cells) > 44 else 0,
                            "camel_slender_gray": safe_float(cells[45]) if len(cells) > 45 else 0,
                            "dp_camel_slender_grayline": safe_float(cells[46]) if len(cells) > 46 else 0,
                            "winston_red_long": safe_float(cells[47]) if len(cells) > 47 else 0,
                            "winston_red": safe_float(cells[48]) if len(cells) > 48 else 0,
                            "winston_blue_long": safe_float(cells[49]) if len(cells) > 49 else 0,
                            "winston_blue": safe_float(cells[50]) if len(cells) > 50 else 0,
                            "winston_gray": safe_float(cells[51]) if len(cells) > 51 else 0,
                            "winston_slims_blue": safe_float(cells[52]) if len(cells) > 52 else 0,
                            "winston_slims_gray": safe_float(cells[53]) if len(cells) > 53 else 0,
                            "winston_slims_q_line": safe_float(cells[54]) if len(cells) > 54 else 0,
                            "winston_xsence_black": safe_float(cells[55]) if len(cells) > 55 else 0,
                            "winston_xsence_gray": safe_float(cells[56]) if len(cells) > 56 else 0,
                            "winston_dark_blue_long": safe_float(cells[57]) if len(cells) > 57 else 0,
                            "winston_dark_blue": safe_float(cells[58]) if len(cells) > 58 else 0,
                            "winston_deep_blue": safe_float(cells[59]) if len(cells) > 59 else 0,
                            "winston_slender_blue_long": safe_float(cells[60]) if len(cells) > 60 else 0,
                            "winston_slender_blue": safe_float(cells[61]) if len(cells) > 61 else 0,
                            "winston_slender_gray": safe_float(cells[62]) if len(cells) > 62 else 0,
                            "winston_slender_dark_blue": safe_float(cells[63]) if len(cells) > 63 else 0,
                            "winston_slender_q_line": safe_float(cells[64]) if len(cells) > 64 else 0,
                            "monte_carlo_red": safe_float(cells[65]) if len(cells) > 65 else 0,
                            "monte_carlo_dark_blue_long": safe_float(cells[66]) if len(cells) > 66 else 0,
                            "monte_carlo_dark_blue": safe_float(cells[67]) if len(cells) > 67 else 0,
                            "monte_carlo_slender_dark_blue": safe_float(cells[68]) if len(cells) > 68 else 0,
                            "ld_slims": safe_float(cells[69]) if len(cells) > 69 else 0,
                            "ld_blue_long": safe_float(cells[70]) if len(cells) > 70 else 0,
                            "ld_blue": safe_float(cells[71]) if len(cells) > 71 else 0,
                            "toplam_gun_sku": safe_float(cells[72]) if len(cells) > 72 else 0,
                            # Marka Toplamları (73-80)
                            "camel_toplam": safe_float(cells[73]) if len(cells) > 73 else 0,
                            "winston_toplam": safe_float(cells[74]) if len(cells) > 74 else 0,
                            "mcarlo_toplam": safe_float(cells[75]) if len(cells) > 75 else 0,
                            "myo_camel": safe_float(cells[76]) if len(cells) > 76 else 0,
                            "ld_toplam": safe_float(cells[77]) if len(cells) > 77 else 0,
                            "toplam": safe_float(cells[78]) if len(cells) > 78 else 0,
                            "kasa": safe_float(cells[79]) if len(cells) > 79 else 0,
                            "hedef_das": safe_float(cells[80]) if len(cells) > 80 else 0,
                            # Gerçekleşen (81-88)
                            "camel_gerc": safe_float(cells[81]) if len(cells) > 81 else 0,
                            "winston_gerc": safe_float(cells[82]) if len(cells) > 82 else 0,
                            "mcarlo_gerc": safe_float(cells[83]) if len(cells) > 83 else 0,
                            "myo_camel_gerc": safe_float(cells[84]) if len(cells) > 84 else 0,
                            "ld_gerc": safe_float(cells[85]) if len(cells) > 85 else 0,
                            "toplam_gerc": safe_float(cells[86]) if len(cells) > 86 else 0,
                            "kasa_gerc": safe_float(cells[87]) if len(cells) > 87 else 0,
                            "gerc_das": safe_float(cells[88]) if len(cells) > 88 else 0,
                            # Kanal Bazlı (89-99)
                            "bak_01": safe_float(cells[89]) if len(cells) > 89 else 0,
                            "mar_02": safe_float(cells[90]) if len(cells) > 90 else 0,
                            "bfe_03": safe_float(cells[91]) if len(cells) > 91 else 0,
                            "kye_04": safe_float(cells[92]) if len(cells) > 92 else 0,
                            "tek_05": safe_float(cells[93]) if len(cells) > 93 else 0,
                            "ben_07": safe_float(cells[94]) if len(cells) > 94 else 0,
                            "ask_08": safe_float(cells[95]) if len(cells) > 95 else 0,
                            "czv_11": safe_float(cells[96]) if len(cells) > 96 else 0,
                            "yznc_12": safe_float(cells[97]) if len(cells) > 97 else 0,
                            "tut_14": safe_float(cells[98]) if len(cells) > 98 else 0,
                            "tus_15": safe_float(cells[99]) if len(cells) > 99 else 0,
                            # Stand Sayıları (100-103)
                            "jti": safe_float(cells[100]) if len(cells) > 100 else 0,
                            "pmi": safe_float(cells[101]) if len(cells) > 101 else 0,
                            "bat": safe_float(cells[102]) if len(cells) > 102 else 0,
                            "rut_say": safe_float(cells[103]) if len(cells) > 103 else 0,
                            # Yıllık SKU (104-115)
                            "w_dark_blue_ks": safe_float(cells[104]) if len(cells) > 104 else 0,
                            "w_slender_blue_ks": safe_float(cells[105]) if len(cells) > 105 else 0,
                            "w_dark_blue_long": safe_float(cells[106]) if len(cells) > 106 else 0,
                            "mcarlo_slender_dark_blue_yil": safe_float(cells[107]) if len(cells) > 107 else 0,
                            "w_slim_blue": safe_float(cells[108]) if len(cells) > 108 else 0,
                            "w_blue_ks": safe_float(cells[109]) if len(cells) > 109 else 0,
                            "w_slender_blue_long": safe_float(cells[110]) if len(cells) > 110 else 0,
                            "camel_slender_blue_yil": safe_float(cells[111]) if len(cells) > 111 else 0,
                            "mcarlo_dark_blue_ks": safe_float(cells[112]) if len(cells) > 112 else 0,
                            "mcarlo_dark_blue_long_yil": safe_float(cells[113]) if len(cells) > 113 else 0,
                            "w_slender_q_line_2025": safe_float(cells[114]) if len(cells) > 114 else 0,
                            "w_slender_q_line_2026": safe_float(cells[115]) if len(cells) > 115 else 0,
                            "frekans_ort": safe_float(cells[116]) if len(cells) > 116 else 0,
                        }
                        dst_data_list.append(dst_record)
            
            if dst_data_list:
                await db.dst_data.insert_many(dst_data_list)
                logger.info(f"Inserted {len(dst_data_list)} DST data records")
            
            # Process DSM Teams (TEAM-I row 21, TEAM-II row 11)
            logger.info("Processing DSM Teams...")
            
            # TEAM-II (satır 11 - index 10)
            team2_row = rows[10]
            team2_cells = [cell.v for cell in team2_row]
            
            # TEAM-I (satır 21 - index 20)
            team1_row = rows[20]
            team1_cells = [cell.v for cell in team1_row]
            
            def create_team_data(cells, team_name, dsm_name, dst_list):
                return {
                    "team_name": team_name,
                    "dsm_name": dsm_name,
                    "dst_list": dst_list,
                    "bayi_sayisi": safe_float(cells[1]) if len(cells) > 1 else 0,
                    "aktif_bayi_sayisi": safe_float(cells[2]) if len(cells) > 2 else 0,
                    "pasif_bayi_sayisi": safe_float(cells[3]) if len(cells) > 3 else 0,
                    "aralik_hedef": safe_float(cells[4]) if len(cells) > 4 else 0,
                    "aralik_satis": safe_float(cells[5]) if len(cells) > 5 else 0,
                    "kalan_satis": safe_float(cells[6]) if len(cells) > 6 else 0,
                    "hedef_basari_orani": safe_float(cells[7]) if len(cells) > 7 else 0,
                    "tahsilat_hedef": safe_float(cells[8]) if len(cells) > 8 else 0,
                    "tahsilat_tutari": safe_float(cells[9]) if len(cells) > 9 else 0,
                    "ay_hedef_ziyaret": safe_float(cells[10]) if len(cells) > 10 else 0,
                    "ziyaret_gerceklesen": safe_float(cells[11]) if len(cells) > 11 else 0,
                    "drop_rate": safe_float(cells[12]) if len(cells) > 12 else 0,
                    "basarili_satis": safe_float(cells[13]) if len(cells) > 13 else 0,
                    "basarili_satis_yuzde": safe_float(cells[14]) if len(cells) > 14 else 0,
                    "carili_bayi_sayisi": safe_float(cells[15]) if len(cells) > 15 else 0,
                    "gun_0": safe_float(cells[16]) if len(cells) > 16 else 0,
                    "gun_1": safe_float(cells[17]) if len(cells) > 17 else 0,
                    "gun_2": safe_float(cells[18]) if len(cells) > 18 else 0,
                    "gun_3": safe_float(cells[19]) if len(cells) > 19 else 0,
                    "gun_4": safe_float(cells[20]) if len(cells) > 20 else 0,
                    "gun_5": safe_float(cells[21]) if len(cells) > 21 else 0,
                    "gun_6": safe_float(cells[22]) if len(cells) > 22 else 0,
                    "gun_7": safe_float(cells[23]) if len(cells) > 23 else 0,
                    "gun_8": safe_float(cells[24]) if len(cells) > 24 else 0,
                    "gun_9": safe_float(cells[25]) if len(cells) > 25 else 0,
                    "gun_10": safe_float(cells[26]) if len(cells) > 26 else 0,
                    "gun_11": safe_float(cells[27]) if len(cells) > 27 else 0,
                    "gun_12": safe_float(cells[28]) if len(cells) > 28 else 0,
                    "gun_13": safe_float(cells[29]) if len(cells) > 29 else 0,
                    "gun_14_uzeri": safe_float(cells[30]) if len(cells) > 30 else 0,
                    "cari_toplam": safe_float(cells[31]) if len(cells) > 31 else 0,
                    "toplam_gun_sku": safe_float(cells[72]) if len(cells) > 72 else 0,
                    # Hedefler
                    "camel_toplam": safe_float(cells[73]) if len(cells) > 73 else 0,
                    "winston_toplam": safe_float(cells[74]) if len(cells) > 74 else 0,
                    "mcarlo_toplam": safe_float(cells[75]) if len(cells) > 75 else 0,
                    "myo_camel": safe_float(cells[76]) if len(cells) > 76 else 0,
                    "ld_toplam": safe_float(cells[77]) if len(cells) > 77 else 0,
                    "toplam": safe_float(cells[78]) if len(cells) > 78 else 0,
                    "kasa": safe_float(cells[79]) if len(cells) > 79 else 0,
                    "hedef_das": safe_float(cells[80]) if len(cells) > 80 else 0,
                    # Satışlar
                    "camel_gerc": safe_float(cells[81]) if len(cells) > 81 else 0,
                    "winston_gerc": safe_float(cells[82]) if len(cells) > 82 else 0,
                    "mcarlo_gerc": safe_float(cells[83]) if len(cells) > 83 else 0,
                    "myo_camel_gerc": safe_float(cells[84]) if len(cells) > 84 else 0,
                    "ld_gerc": safe_float(cells[85]) if len(cells) > 85 else 0,
                    "toplam_gerc": safe_float(cells[86]) if len(cells) > 86 else 0,
                    "kasa_gerc": safe_float(cells[87]) if len(cells) > 87 else 0,
                    "gerc_das": safe_float(cells[88]) if len(cells) > 88 else 0,
                    # Bayi Tipleri
                    "bak_01": safe_float(cells[89]) if len(cells) > 89 else 0,
                    "mar_02": safe_float(cells[90]) if len(cells) > 90 else 0,
                    "bfe_03": safe_float(cells[91]) if len(cells) > 91 else 0,
                    "kye_04": safe_float(cells[92]) if len(cells) > 92 else 0,
                    "tek_05": safe_float(cells[93]) if len(cells) > 93 else 0,
                    "ben_07": safe_float(cells[94]) if len(cells) > 94 else 0,
                    "ask_08": safe_float(cells[95]) if len(cells) > 95 else 0,
                    "czv_11": safe_float(cells[96]) if len(cells) > 96 else 0,
                    "yznc_12": safe_float(cells[97]) if len(cells) > 97 else 0,
                    "tut_14": safe_float(cells[98]) if len(cells) > 98 else 0,
                    "tus_15": safe_float(cells[99]) if len(cells) > 99 else 0,
                    "jti": safe_float(cells[100]) if len(cells) > 100 else 0,
                    "pmi": safe_float(cells[101]) if len(cells) > 101 else 0,
                    "bat": safe_float(cells[102]) if len(cells) > 102 else 0,
                    "rut_say": safe_float(cells[103]) if len(cells) > 103 else 0,
                    # İlk 10 SKU
                    "w_dark_blue_ks": safe_float(cells[104]) if len(cells) > 104 else 0,
                    "w_slender_blue_ks": safe_float(cells[105]) if len(cells) > 105 else 0,
                    "w_dark_blue_long": safe_float(cells[106]) if len(cells) > 106 else 0,
                    "mcarlo_slender_dark_blue_yil": safe_float(cells[107]) if len(cells) > 107 else 0,
                    "w_slim_blue": safe_float(cells[108]) if len(cells) > 108 else 0,
                    "w_blue_ks": safe_float(cells[109]) if len(cells) > 109 else 0,
                    "w_slender_blue_long": safe_float(cells[110]) if len(cells) > 110 else 0,
                    "camel_slender_blue_yil": safe_float(cells[111]) if len(cells) > 111 else 0,
                    "mcarlo_dark_blue_ks": safe_float(cells[112]) if len(cells) > 112 else 0,
                    "mcarlo_dark_blue_long_yil": safe_float(cells[113]) if len(cells) > 113 else 0,
                    "w_slender_q_line_2025": safe_float(cells[114]) if len(cells) > 114 else 0,
                    "w_slender_q_line_2026": safe_float(cells[115]) if len(cells) > 115 else 0,
                    "frekans_ort": safe_float(cells[116]) if len(cells) > 116 else 0,
                }
            
            team2_data = create_team_data(
                team2_cells,
                "TEAM-II",
                "MURAT YÖRÜKOĞLU",
                ["KEMAL BANİ", "COŞKUN ÇİMEN", "MUSTAFA KAĞAN KAYA", "MUSTAFA HARMANCI", 
                 "KAZIM KARABEKİR ÖRAN", "TUNAHAN IŞILAK", "MEVLÜT ŞEKER", "TAHİR UÇAR", "YASİN TUĞRA DAĞLI"]
            )
            
            team1_data = create_team_data(
                team1_cells,
                "TEAM-I",
                "OSMAN DİNÇOL",
                ["HÜSEYİN AYHAN AKMAN", "MUSTAFA USLU", "HASAN ALİ AKDAĞ", "AHMET GÖKMEN",
                 "LÜTFİ UYSAL", "ŞERAFETTİN BÜYÜKTAŞDELEN", "BURAK KÜÇÜKŞANTÜRK", "YASİN AVCI", "MUSTAFA İBİŞ"]
            )
            
            await db.dsm_teams.insert_many([team1_data, team2_data])
            logger.info("Inserted 2 DSM team records")
            
            # Process TTE Data (rows 24-27 for TTE info, rows 28-32 for stand info)
            logger.info("Processing TTE Data...")
            tte_data_list = []
            
            for i in range(23, 27):  # Rows 24-27 (index 23-26)
                row = rows[i]
                cells = [cell.v for cell in row]
                if cells[0]:
                    tte_name = safe_str(cells[0])
                    tte_record = {
                        "tte_name": tte_name,
                        "bayi_sayisi": safe_float(cells[1]) if len(cells) > 1 else 0,
                        "aktif_bayi_sayisi": safe_float(cells[2]) if len(cells) > 2 else 0,
                        "pasif_bayi_sayisi": safe_float(cells[3]) if len(cells) > 3 else 0,
                    }
                    tte_data_list.append(tte_record)
            
            # Stand info from rows 29-32 (index 28-31)
            for i, tte_data in enumerate(tte_data_list):
                stand_row_idx = 28 + i  # Rows 29, 30, 31, 32
                if stand_row_idx < len(rows):
                    stand_row = rows[stand_row_idx]
                    stand_cells = [cell.v for cell in stand_row]
                    
                    tte_data["jti"] = safe_float(stand_cells[1]) if len(stand_cells) > 1 else 0
                    tte_data["jti_stand"] = safe_float(stand_cells[2]) if len(stand_cells) > 2 else 0
                    tte_data["pmi"] = safe_float(stand_cells[3]) if len(stand_cells) > 3 else 0
                    tte_data["pmi_stand"] = safe_float(stand_cells[4]) if len(stand_cells) > 4 else 0
                    tte_data["bat"] = safe_float(stand_cells[5]) if len(stand_cells) > 5 else 0
                    tte_data["bat_stand"] = safe_float(stand_cells[6]) if len(stand_cells) > 6 else 0
                    # Bayi sınıfları (H-N sütunları, index 7-13)
                    tte_data["sinif_a"] = safe_float(stand_cells[7]) if len(stand_cells) > 7 else 0
                    tte_data["sinif_a_plus"] = safe_float(stand_cells[8]) if len(stand_cells) > 8 else 0
                    tte_data["sinif_b"] = safe_float(stand_cells[9]) if len(stand_cells) > 9 else 0
                    tte_data["sinif_c"] = safe_float(stand_cells[10]) if len(stand_cells) > 10 else 0
                    tte_data["sinif_d"] = safe_float(stand_cells[11]) if len(stand_cells) > 11 else 0
                    tte_data["sinif_e"] = safe_float(stand_cells[12]) if len(stand_cells) > 12 else 0
                    tte_data["sinif_e_minus"] = safe_float(stand_cells[13]) if len(stand_cells) > 13 else 0
            
            if tte_data_list:
                await db.tte_data.insert_many(tte_data_list)
                logger.info(f"Inserted {len(tte_data_list)} TTE data records")
            
            # Process Distributor Totals from Row 22
            logger.info("Processing Distributor Totals...")
            await db.distributor_totals.delete_many({})
            
            row22 = rows[21]  # Row 22 (0-indexed = 21)
            cells = [cell.v for cell in row22]
            
            totals = {
                "type": "totals",
                # B22-AH22 (columns 1-33)
                "bayi_sayisi": safe_float(cells[1]) if len(cells) > 1 else 0,
                "aktif_bayi_sayisi": safe_float(cells[2]) if len(cells) > 2 else 0,
                "pasif_bayi_sayisi": safe_float(cells[3]) if len(cells) > 3 else 0,
                "aralik_hedef": safe_float(cells[4]) if len(cells) > 4 else 0,
                "aralik_satis": safe_float(cells[5]) if len(cells) > 5 else 0,
                "kalan_satis": safe_float(cells[6]) if len(cells) > 6 else 0,
                "hedef_basari_orani": safe_float(cells[7]) if len(cells) > 7 else 0,
                "tahsilat_hedef": safe_float(cells[8]) if len(cells) > 8 else 0,
                "tahsilat_tutari": safe_float(cells[9]) if len(cells) > 9 else 0,
                "ay_hedef_ziyaret": safe_float(cells[10]) if len(cells) > 10 else 0,
                "ziyaret_gerceklesen": safe_float(cells[11]) if len(cells) > 11 else 0,
                "drop_rate": safe_float(cells[12]) if len(cells) > 12 else 0,
                "basarili_satis": safe_float(cells[13]) if len(cells) > 13 else 0,
                "basarili_satis_yuzde": safe_float(cells[14]) if len(cells) > 14 else 0,
                "carili_bayi_sayisi": safe_float(cells[15]) if len(cells) > 15 else 0,
                "gun_0": safe_float(cells[16]) if len(cells) > 16 else 0,
                "gun_1": safe_float(cells[17]) if len(cells) > 17 else 0,
                "gun_2": safe_float(cells[18]) if len(cells) > 18 else 0,
                "gun_3": safe_float(cells[19]) if len(cells) > 19 else 0,
                "gun_4": safe_float(cells[20]) if len(cells) > 20 else 0,
                "gun_5": safe_float(cells[21]) if len(cells) > 21 else 0,
                "gun_6": safe_float(cells[22]) if len(cells) > 22 else 0,
                "gun_7": safe_float(cells[23]) if len(cells) > 23 else 0,
                "gun_8": safe_float(cells[24]) if len(cells) > 24 else 0,
                "gun_9": safe_float(cells[25]) if len(cells) > 25 else 0,
                "gun_10": safe_float(cells[26]) if len(cells) > 26 else 0,
                "gun_11": safe_float(cells[27]) if len(cells) > 27 else 0,
                "gun_12": safe_float(cells[28]) if len(cells) > 28 else 0,
                "gun_13": safe_float(cells[29]) if len(cells) > 29 else 0,
                "gun_14_uzeri": safe_float(cells[30]) if len(cells) > 30 else 0,
                "cari_toplam": safe_float(cells[31]) if len(cells) > 31 else 0,
                "loy_verilen_bayi_sayisi": safe_float(cells[32]) if len(cells) > 32 else 0,
                "loy_bayi_mahsuplasma_tutari": safe_float(cells[33]) if len(cells) > 33 else 0,
                # BV22-CZ22 (columns 73-103)
                "camel_hedef": safe_float(cells[73]) if len(cells) > 73 else 0,
                "winston_hedef": safe_float(cells[74]) if len(cells) > 74 else 0,
                "mcarlo_hedef": safe_float(cells[75]) if len(cells) > 75 else 0,
                "myo_camel_hedef": safe_float(cells[76]) if len(cells) > 76 else 0,
                "ld_hedef": safe_float(cells[77]) if len(cells) > 77 else 0,
                "toplam_hedef": safe_float(cells[78]) if len(cells) > 78 else 0,
                "kasa_hedef": safe_float(cells[79]) if len(cells) > 79 else 0,
                "hedef_das": safe_float(cells[80]) if len(cells) > 80 else 0,
                "camel_satis": safe_float(cells[81]) if len(cells) > 81 else 0,
                "winston_satis": safe_float(cells[82]) if len(cells) > 82 else 0,
                "mcarlo_satis": safe_float(cells[83]) if len(cells) > 83 else 0,
                "myo_camel_satis": safe_float(cells[84]) if len(cells) > 84 else 0,
                "ld_satis": safe_float(cells[85]) if len(cells) > 85 else 0,
                "toplam_satis": safe_float(cells[86]) if len(cells) > 86 else 0,
                "kasa_satis": safe_float(cells[87]) if len(cells) > 87 else 0,
                "gerc_das": safe_float(cells[88]) if len(cells) > 88 else 0,
                "bak_01": safe_float(cells[89]) if len(cells) > 89 else 0,
                "mar_02": safe_float(cells[90]) if len(cells) > 90 else 0,
                "bfe_03": safe_float(cells[91]) if len(cells) > 91 else 0,
                "kye_04": safe_float(cells[92]) if len(cells) > 92 else 0,
                "tek_05": safe_float(cells[93]) if len(cells) > 93 else 0,
                "ben_07": safe_float(cells[94]) if len(cells) > 94 else 0,
                "ask_08": safe_float(cells[95]) if len(cells) > 95 else 0,
                "czv_11": safe_float(cells[96]) if len(cells) > 96 else 0,
                "yznc_12": safe_float(cells[97]) if len(cells) > 97 else 0,
                "tut_14": safe_float(cells[98]) if len(cells) > 98 else 0,
                "tus_15": safe_float(cells[99]) if len(cells) > 99 else 0,
                "jti": safe_float(cells[100]) if len(cells) > 100 else 0,
                "pmi": safe_float(cells[101]) if len(cells) > 101 else 0,
                "bat": safe_float(cells[102]) if len(cells) > 102 else 0,
                "rut_say": safe_float(cells[103]) if len(cells) > 103 else 0,
                # DL22-DS22 (columns 115-116)
                "qline_2026_satis": safe_float(cells[115]) if len(cells) > 115 else 0,
                "frekans_ort": safe_float(cells[116]) if len(cells) > 116 else 0,
                # DT22 - Q Line Hedef (column 123)
                "qline_hedef": safe_float(cells[123]) if len(cells) > 123 else 0,
            }
            
            await db.distributor_totals.insert_one(totals)
            logger.info("Inserted distributor totals")
            
            # Process Günlük Ekip Raporu Verileri
            logger.info("Processing Günlük Ekip Raporu Verileri...")
            await db.ekip_raporu.delete_many({})
            
            sku_fields = [
                "skt_camel_yellow_100", "camel_brown", "camel_black", "camel_white",
                "camel_yellow_sp", "camel_yellow", "camel_deep_blue_long", "camel_deep_blue",
                "camel_yellow_long", "camel_slender_blue", "dp_camel_slender_blueline",
                "camel_slender_gray", "dp_camel_slender_grayline", "winston_red_long",
                "winston_red", "winston_blue_long", "winston_blue", "winston_gray",
                "winston_slims_blue", "winston_slims_gray", "winston_slims_q_line",
                "winston_xsence_black", "winston_xsence_gray", "winston_dark_blue_long",
                "winston_dark_blue", "winston_deep_blue", "winston_slender_blue_long",
                "winston_slender_blue", "winston_slender_gray", "winston_slender_dark_blue",
                "winston_slender_q_line", "monte_carlo_red", "monte_carlo_dark_blue_long",
                "monte_carlo_dark_blue", "monte_carlo_slender_dark_blue", "ld_slims",
                "ld_blue_long", "ld_blue", "toplam"
            ]
            
            with wb.get_sheet("Günlük Ekip Raporu Verileri.") as sheet:
                ekip_rows = list(sheet.rows())
                ekip_data = []
                yil_toplam_karton = None
                yil_toplam_kasa = None
                
                for row in ekip_rows[1:]:
                    cells = [c.v for c in row]
                    if len(cells) > 1:
                        ay = safe_str(cells[0])
                        tarih_raw = cells[1]
                        
                        # Skip summary rows
                        if isinstance(tarih_raw, str):
                            if "TOPLAM" in tarih_raw.upper():
                                if "YIL TOPLAM KARTON" in tarih_raw.upper():
                                    yil_toplam_karton = {f: safe_float(cells[i+2]) for i, f in enumerate(sku_fields) if i+2 < len(cells)}
                                elif "YIL TOPLAM KASA" in tarih_raw.upper():
                                    yil_toplam_kasa = {f: safe_float(cells[i+2]) for i, f in enumerate(sku_fields) if i+2 < len(cells)}
                            continue
                        
                        if tarih_raw:
                            record = {
                                "ay": ay,
                                "tarih": tarih_raw,
                            }
                            for i, field in enumerate(sku_fields):
                                record[field] = safe_float(cells[i+2]) if i+2 < len(cells) else 0
                            ekip_data.append(record)
                
                if ekip_data:
                    await db.ekip_raporu.insert_many(ekip_data)
                    logger.info(f"Inserted {len(ekip_data)} ekip raporu records")
                
                # Save yearly totals
                if yil_toplam_karton or yil_toplam_kasa:
                    await db.ekip_raporu_toplam.delete_many({})
                    toplam_doc = {
                        "yil_toplam_karton": yil_toplam_karton or {},
                        "yil_toplam_kasa": yil_toplam_kasa or {}
                    }
                    await db.ekip_raporu_toplam.insert_one(toplam_doc)
            
            # Process STİL AY SATIŞ
            logger.info("Processing STİL AY SATIŞ...")
            await db.stil_ay_satis.delete_many({})
            
            with wb.get_sheet("STİL AY SATIŞ") as sheet:
                stil_rows = list(sheet.rows())
                stil_data = []
                
                for row in stil_rows[1:]:
                    cells = [c.v for c in row]
                    if len(cells) > 0 and cells[0]:
                        record = {
                            "ay": safe_str(cells[0]),
                        }
                        for i, field in enumerate(sku_fields):
                            record[field] = safe_float(cells[i+1]) if i+1 < len(cells) else 0
                        stil_data.append(record)
                
                if stil_data:
                    await db.stil_ay_satis.insert_many(stil_data)
                    logger.info(f"Inserted {len(stil_data)} stil ay satis records")
            
            # Process PERSONEL DATA
            logger.info("Processing PERSONEL DATA...")
            await db.personel_data.delete_many({})
            
            with wb.get_sheet("PERSONEL DATA") as sheet:
                personel_rows = list(sheet.rows())
                personel_data = []
                
                for row in personel_rows[2:]:  # Skip header rows
                    cells = [c.v for c in row]
                    if len(cells) > 3 and cells[3]:  # ADI field
                        record = {
                            "sira_no": safe_float(cells[0]),
                            "bolge": safe_str(cells[1]),
                            "distributor": safe_str(cells[2]),
                            "adi": safe_str(cells[3]),
                            "pozisyonu": safe_str(cells[4]),
                            "cep_telefonu": safe_str(cells[5]),
                            "yakini": safe_str(cells[6]),
                            "yakini_telefon": safe_str(cells[7]),
                            "kan_grubu": safe_str(cells[8]),
                            "src": safe_str(cells[9]),
                            "src_verilis": cells[10],
                            "psikoteknik_verilis": cells[11],
                            "psikoteknik_gecerlilik": cells[12],
                            "mezuniyet": safe_str(cells[13]),
                            "bolum": safe_str(cells[14]),
                            "arac_plaka": safe_str(cells[15]) if len(cells) > 15 else None
                        }
                        personel_data.append(record)
                
                if personel_data:
                    await db.personel_data.insert_many(personel_data)
                    logger.info(f"Inserted {len(personel_data)} personel data records")
    
    # Process RUT sayfası - Ayrı bir workbook açışı ile
    logger.info("Processing RUT...")
    await db.rut_data.delete_many({})
    
    try:
        with pyxlsb.open_workbook(file_path) as wb_rut:
            with wb_rut.get_sheet("RUT") as sheet:
                rut_rows = list(sheet.rows())
                rut_data = []
                
                for row in rut_rows[1:]:  # Skip header
                    cells = [c.v for c in row]
                    if len(cells) > 6 and cells[5]:  # MusteriKod required
                        rut_aciklama = safe_str(cells[3]) if len(cells) > 3 else ""
                        
                        # RutAciklama'dan DST adı ve gün çıkar
                        dst_name = ""
                        gun = ""
                        if rut_aciklama:
                            # "KEMAL BANİ Pazartesi" formatında
                            # Önce uzun günlerden başla (Pazartesi -> Pazar'dan önce kontrol edilmeli)
                            gunler = ["Pazartesi", "Cumartesi", "Perşembe", "Çarşamba", "Salı", "Cuma", "Pazar"]
                            for g in gunler:
                                # Tam kelime olarak ara
                                if rut_aciklama.endswith(g) or f" {g}" in rut_aciklama:
                                    gun = g
                                    # Son kelime olarak günü çıkar
                                    idx = rut_aciklama.rfind(g)
                                    if idx > 0:
                                        dst_name = rut_aciklama[:idx].strip()
                                    break
                        
                        record = {
                            "dist_kod": safe_str(cells[0]) if len(cells) > 0 else "",
                            "dist_unvan": safe_str(cells[1]) if len(cells) > 1 else "",
                            "rut_kod": safe_str(cells[2]) if len(cells) > 2 else "",
                            "rut_aciklama": rut_aciklama,
                            "dst_name": dst_name,
                            "gun": gun,
                            "ziyaret_sira": int(safe_float(cells[4])) if len(cells) > 4 else 0,
                            "musteri_kod": safe_str(cells[5]) if len(cells) > 5 else "",
                            "musteri_unvan": safe_str(cells[6]) if len(cells) > 6 else "",
                            "musteri_durum": safe_str(cells[7]) if len(cells) > 7 else "",
                            "musteri_grup_kod": safe_str(cells[8]) if len(cells) > 8 else "",
                            "musteri_grup": safe_str(cells[9]) if len(cells) > 9 else "",
                            "musteri_ek_grup": safe_str(cells[10]) if len(cells) > 10 else "",
                            "adres": safe_str(cells[11]) if len(cells) > 11 else "",
                        }
                        rut_data.append(record)
                
                if rut_data:
                    await db.rut_data.insert_many(rut_data)
                    logger.info(f"Inserted {len(rut_data)} rut records")
    except Exception as e:
        logger.warning(f"Could not process RUT sheet: {e}")
    
    # Process Bayi Hedef sheet
    try:
        logger.info("Processing Bayi Hedef...")
        with pyxlsb.open_workbook(file_path) as wb_hedef:
            with wb_hedef.get_sheet('Bayi Hedef') as sheet:
                rows = list(sheet.rows())
                bayi_hedef_data = []
                
                for row in rows[1:]:  # Skip header
                    cells = [cell.v for cell in row]
                    if len(cells) > 20 and cells[1]:
                        bayi_kodu = str(int(cells[1])) if isinstance(cells[1], float) else str(cells[1])
                        bayi_hedef = {
                            "bayi_kodu": bayi_kodu,
                        "bayi_adi": safe_str(cells[2]) if len(cells) > 2 else "",
                        "dst": safe_str(cells[3]) if len(cells) > 3 else "",
                        "sinif": safe_str(cells[4]) if len(cells) > 4 else "",
                        "camel_hedef": safe_float(cells[6]) if len(cells) > 6 else 0,
                        "winston_hedef": safe_float(cells[7]) if len(cells) > 7 else 0,
                        "mcarlo_hedef": safe_float(cells[8]) if len(cells) > 8 else 0,
                        "myo_camel_hedef": safe_float(cells[9]) if len(cells) > 9 else 0,
                        "ld_hedef": safe_float(cells[10]) if len(cells) > 10 else 0,
                        "ay_toplam_hedef": safe_float(cells[11]) if len(cells) > 11 else 0,
                        "camel_satis": safe_float(cells[15]) if len(cells) > 15 else 0,
                        "winston_satis": safe_float(cells[16]) if len(cells) > 16 else 0,
                        "mcarlo_satis": safe_float(cells[17]) if len(cells) > 17 else 0,
                        "myo_camel_satis": safe_float(cells[18]) if len(cells) > 18 else 0,
                        "ld_satis": safe_float(cells[19]) if len(cells) > 19 else 0,
                        "ay_toplam_satis": safe_float(cells[20]) if len(cells) > 20 else 0,
                    }
                    bayi_hedef_data.append(bayi_hedef)
            
            if bayi_hedef_data:
                await db.bayi_hedef.delete_many({})
                await db.bayi_hedef.insert_many(bayi_hedef_data)
                logger.info(f"Inserted {len(bayi_hedef_data)} bayi_hedef records")
    except Exception as e:
        logger.warning(f"Could not process Bayi Hedef sheet: {e}")
    
    # Process Fatura Eki (Loyalty bayileri)
    try:
        logger.info("Processing FATURA EKİ (Loyalty)...")
        with pyxlsb.open_workbook(file_path) as wb_loyalty:
            with wb_loyalty.get_sheet('FATURA EKİ') as sheet:
                rows = list(sheet.rows())
                loyalty_data = []
                
                for row in rows[4:]:  # Start from row 5 (index 4)
                    cells = [cell.v for cell in row]
                    if len(cells) > 17 and cells[5]:  # F column - bayi adı
                        bayi_kodu = str(int(cells[4])) if isinstance(cells[4], float) else str(cells[4]) if cells[4] else ""
                        loyalty = {
                            "bayi_kodu": bayi_kodu,
                            "bayi_adi": safe_str(cells[5]) if len(cells) > 5 else "",
                            "durum": safe_str(cells[6]) if len(cells) > 6 else "",
                            "dsm": safe_str(cells[7]) if len(cells) > 7 else "",
                            "tte": safe_str(cells[8]) if len(cells) > 8 else "",
                            "dst": safe_str(cells[9]) if len(cells) > 9 else "",
                            "kanal": safe_str(cells[10]) if len(cells) > 10 else "",
                            "kod": safe_str(cells[11]) if len(cells) > 11 else "",
                            "sinif": safe_str(cells[12]) if len(cells) > 12 else "",
                            "stand_tipi": safe_str(cells[13]) if len(cells) > 13 else "",
                            "sozlesme_no": safe_str(cells[15]) if len(cells) > 15 else "",
                            "odeme_tutari": safe_float(cells[16]) if len(cells) > 16 else 0,
                            "sozlesme_tutari": safe_float(cells[17]) if len(cells) > 17 else 0,
                        }
                        loyalty_data.append(loyalty)
                
                if loyalty_data:
                    await db.loyalty_bayiler.delete_many({})
                    await db.loyalty_bayiler.insert_many(loyalty_data)
                    logger.info(f"Inserted {len(loyalty_data)} loyalty_bayiler records")
    except Exception as e:
        logger.warning(f"Could not process FATURA EKİ sheet: {e}")
    
    # Son güncelleme zamanını kaydet
    from datetime import datetime
    await db.system_info.delete_many({})
    await db.system_info.insert_one({
        "son_guncelleme": datetime.now().isoformat(),
        "type": "excel_upload"
    })
    
    # Create indexes
    await db.bayiler.create_index("bayi_kodu")
    await db.bayiler.create_index("bayi_kodu_ascii")
    await db.bayiler.create_index("bayi_unvani_ascii")
    await db.faturalar.create_index("bayi_kodu")
    await db.faturalar.create_index("matbu_no")
    await db.belge_detay.create_index("matbu_no")
    await db.tahsilatlar.create_index("bayi_kodu")
    await db.konya_gun.create_index("bayi_kodu")
    await db.stand_raporu.create_index("bayi_durumu")
    await db.bayi_hedef.create_index("bayi_kodu")
    await db.loyalty_bayiler.create_index("bayi_kodu")
    
    logger.info("Excel processing completed!")


# Health check endpoint for Kubernetes probes
@app.get("/health")
async def health_check():
    return {"status": "healthy"}   

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
